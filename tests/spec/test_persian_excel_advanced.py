"""Advanced Persian Excel preservation and visualization tests."""

from __future__ import annotations

from pathlib import Path
from typing import Iterator, Optional

import openpyxl
import pytest
from openpyxl.chart import LineChart, Reference

from tests.fixtures.state import CleanupFixtures


def _extract_title_text(title_obj: Optional[object]) -> Optional[str]:
    """Return raw Persian text from openpyxl title structures."""

    if title_obj is None:
        return None
    rich = getattr(getattr(title_obj, "tx", None), "rich", None)
    if rich and getattr(rich, "p", None):
        try:
            return rich.p[0].r[0].t
        except (AttributeError, IndexError):  # pragma: no cover - defensive guard
            return None
    text = getattr(title_obj, "title", None)
    return text if isinstance(text, str) else None


@pytest.fixture(name="excel_adv_state")
def fixture_excel_adv_state(cleanup_fixtures: CleanupFixtures) -> Iterator[CleanupFixtures]:
    """Ensure Redis/metrics state is reset before and after advanced Excel tests."""

    cleanup_fixtures.flush_state()
    yield cleanup_fixtures
    cleanup_fixtures.flush_state()


@pytest.mark.integration
@pytest.mark.excel
@pytest.mark.timeout(20)
def test_excel_formula_preservation(tmp_path: Path, excel_adv_state: CleanupFixtures) -> None:
    """Ensure Persian right-to-left formulas survive workbook round-trips.

    Example:
        >>> # Formulas remain intact after save/load
        >>> path = tmp_path / "نمونه.xlsx"  # doctest: +SKIP
    """

    path = tmp_path / f"{excel_adv_state.namespace}-formulas.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "فرمول‌ها"
    # Force RTL layout to emulate Persian Excel usage.
    sheet.sheet_view.rightToLeft = True
    sheet["A1"] = "۱۲"
    sheet["B1"] = "۳۴"
    sheet["C1"] = "=SUM(A1:B1)"
    workbook.save(path)

    reloaded = openpyxl.load_workbook(path, data_only=False)
    sheet_rtl = reloaded["فرمول‌ها"]
    context = excel_adv_state.context(
        path=str(path),
        formula=str(sheet_rtl["C1"].value),
        rtl=sheet_rtl.sheet_view.rightToLeft,
    )
    assert sheet_rtl.sheet_view.rightToLeft is True, f"جهت برگه باید راست به چپ باشد: {context}"
    assert sheet_rtl["C1"].value == "=SUM(A1:B1)", f"فرمول پس از ذخیره حفظ نشد: {context}"
    assert sheet_rtl["C1"].data_type == "f", f"نوع سلول فرمولی نیست: {context}"


@pytest.mark.integration
@pytest.mark.excel
@pytest.mark.timeout(20)
def test_excel_chart_rendering(tmp_path: Path, excel_adv_state: CleanupFixtures) -> None:
    """Verify Persian-labelled charts persist with correct metadata.

    Example:
        >>> # Executed within pytest to render a Persian chart
        >>> tmp_path = Path('charts')  # doctest: +SKIP
    """

    path = tmp_path / f"{excel_adv_state.namespace}-chart.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "آمار"
    sheet.append(["ماه", "ثبت‌نام"])
    payload = [
        ("فروردین", 120),
        ("اردیبهشت", 135),
        ("خرداد", 128),
        ("تیر", 142),
    ]
    for month, value in payload:
        sheet.append([month, value])

    chart = LineChart()
    chart.title = "روند ثبت‌نام"
    chart.x_axis.title = "ماه"
    chart.y_axis.title = "تعداد"
    # Categories capture Persian month names to confirm Unicode preservation.
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=len(payload) + 1))
    data_ref = Reference(sheet, min_col=2, min_row=1, max_row=len(payload) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    sheet.add_chart(chart, "E2")
    workbook.save(path)

    reloaded = openpyxl.load_workbook(path)
    sheet_chart = reloaded["آمار"]
    titles = []
    for item in sheet_chart._charts:
        titles.append(_extract_title_text(getattr(item, "title", None)) or "<missing>")
    chart_obj = sheet_chart._charts[0]
    x_axis_title = _extract_title_text(getattr(getattr(chart_obj, "x_axis", None), "title", None))
    y_axis_title = _extract_title_text(getattr(getattr(chart_obj, "y_axis", None), "title", None))
    context = excel_adv_state.context(
        path=str(path),
        charts=len(sheet_chart._charts),
        titles=titles,
        x_axis=x_axis_title,
        y_axis=y_axis_title,
    )
    assert sheet_chart._charts, f"هیچ نموداری در فایل یافت نشد: {context}"
    assert "روند ثبت‌نام" in titles, f"عنوان نمودار فارسی گم شده است: {context}"
    assert x_axis_title == "ماه", f"عنوان محور افقی نادرست است: {context}"
    assert y_axis_title == "تعداد", f"عنوان محور عمودی نادرست است: {context}"
