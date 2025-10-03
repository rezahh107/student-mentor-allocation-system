from __future__ import annotations

import csv
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Iterator

from openpyxl import load_workbook

from phase6_import_to_sabt.sanitization import fold_digits, guard_formula, sanitize_phone, sanitize_text
from phase6_import_to_sabt.xlsx.constants import ALLOWED_EXTENSIONS, MAX_UPLOAD_SIZE_BYTES, RISKY_FORMULA_PREFIXES, SENSITIVE_COLUMNS
from phase6_import_to_sabt.xlsx.utils import cleanup_partials, ensure_max_size, normalized_header


@dataclass(slots=True)
class UploadRow:
    values: dict[str, str]


@dataclass(slots=True)
class UploadResult:
    rows: list[UploadRow]
    excel_safety: dict[str, Any]
    row_counts: dict[str, int]
    format: str


class XLSXUploadReader:
    def __init__(self, *, required_columns: Iterable[str] | None = None) -> None:
        self._required = tuple(required_columns or ("school_code",))

    def read(self, path: Path) -> UploadResult:
        cleanup_partials(path.parent)
        ensure_max_size(path, MAX_UPLOAD_SIZE_BYTES)
        ext = path.suffix.lower()
        if ext not in ALLOWED_EXTENSIONS:
            raise ValueError("UPLOAD_FORMAT_NOT_ALLOWED")
        if ext == ".zip":
            inner_path, inner_ext = self._extract_first_member(path)
            try:
                return self._read_inner(inner_path, inner_ext)
            finally:
                inner_path.unlink(missing_ok=True)
        return self._read_inner(path, ext)

    def _extract_first_member(self, archive_path: Path) -> tuple[Path, str]:
        with zipfile.ZipFile(archive_path) as archive:
            names = [name for name in archive.namelist() if not name.endswith("/")]
            if not names:
                raise ValueError("UPLOAD_ZIP_EMPTY")
            member = names[0]
            if Path(member).is_absolute() or ".." in Path(member).parts:
                raise ValueError("UPLOAD_ZIP_TRAVERSAL")
            suffix = Path(member).suffix.lower()
            if suffix not in {".xlsx", ".csv"}:
                raise ValueError("UPLOAD_ZIP_UNSUPPORTED")
            data = archive.read(member)
            temp_path = archive_path.with_suffix(suffix + ".inner")
            temp_path.write_bytes(data)
            return temp_path, suffix

    def _read_inner(self, path: Path, ext: str) -> UploadResult:
        if ext == ".csv":
            rows = list(self._read_csv(path))
            fmt = "csv"
        else:
            rows = list(self._read_xlsx(path))
            fmt = "xlsx"
        if not rows:
            raise ValueError("UPLOAD_EMPTY")
        headers = rows[0].values.keys()
        for required in self._required:
            if required not in headers:
                raise ValueError("UPLOAD_VALIDATION_ERROR")
        excel_safety = {
            "normalized": True,
            "digit_folded": True,
            "formula_guard": True,
            "sensitive_text": list(SENSITIVE_COLUMNS),
        }
        row_counts = {"Sheet_001": len(rows)}
        return UploadResult(rows=rows, excel_safety=excel_safety, row_counts=row_counts, format=fmt)

    def _read_csv(self, path: Path) -> Iterator[UploadRow]:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            rows = list(reader)
        if not rows:
            return iter(())
        header_map = [normalized_header(col) for col in rows[0]]
        for raw in rows[1:]:
            values = {}
            for index, column in enumerate(header_map):
                value = raw[index] if index < len(raw) else ""
                values[column] = self._sanitize_cell(column, value)
            yield UploadRow(values=values)

    def _read_xlsx(self, path: Path) -> Iterator[UploadRow]:
        workbook = load_workbook(filename=path, read_only=True, data_only=False)
        try:
            sheet = workbook.active
            iterator = sheet.iter_rows(values_only=True)
            try:
                headers = next(iterator)
            except StopIteration:
                return iter(())
            normalized_headers = [normalized_header(str(value or "")) for value in headers]
            for row in iterator:
                values = {}
                for index, column in enumerate(normalized_headers):
                    cell_value = row[index] if index < len(row) else ""
                    values[column] = self._sanitize_cell(column, cell_value)
                yield UploadRow(values=values)
        finally:
            workbook.close()

    def _sanitize_cell(self, column: str, value: Any) -> str:
        if value is None:
            text = ""
        else:
            text = str(value)
        text = sanitize_text(text)
        if column in SENSITIVE_COLUMNS:
            text = fold_digits(text)
            if column == "school_code" and text:
                digits = "".join(ch for ch in text if ch.isdigit())
                if digits:
                    text = digits.zfill(6)
        if column == "mobile":
            text = sanitize_phone(text)
        if text and text[0] in RISKY_FORMULA_PREFIXES:
            text = guard_formula(text)
        return text
