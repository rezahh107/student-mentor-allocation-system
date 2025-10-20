from __future__ import annotations

import csv
import json
from dataclasses import replace
from datetime import datetime, timezone
from pathlib import Path

from sma.phase6_import_to_sabt.models import ExportFilters, ExportOptions, ExportSnapshot

from tests.export.helpers import build_exporter, make_row


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader)


def test_stable_sort_order(tmp_path: Path) -> None:
    base_time = datetime(2024, 1, 1, 9, 0, tzinfo=timezone.utc)
    rows = [
        make_row(idx=3, center=2, group_code=20, school_code=None),
        replace(
            make_row(idx=2, center=0, group_code=5),
            national_id="۰۰۰۰۰۰۰۰۰۲",
            first_name="\u200cاسم",
            school_code=654321,
        ),
        replace(
            make_row(idx=1, center=0, group_code=5, gender=1),
            national_id="0000000001",
            school_code=123456,
        ),
    ]
    exporter = build_exporter(tmp_path, rows)
    filters = ExportFilters(year=1402, center=None)
    snapshot = ExportSnapshot(marker="snap", created_at=base_time)
    options = ExportOptions(chunk_size=2, include_bom=True, output_format="csv")

    manifest = exporter.run(filters=filters, options=options, snapshot=snapshot, clock_now=base_time)

    csv_files = sorted(tmp_path.glob("*.csv"))
    assert len(csv_files) == 2
    assert all(file.name.endswith(".csv") for file in csv_files)

    ordered_records: list[str] = []
    for file in csv_files:
        content = file.read_bytes()
        assert b"\r\n" in content  # CRLF preserved
        rows_data = _read_csv(file)
        for record in rows_data:
            ordered_records.append(record["national_id"])
            assert all("\u200c" not in value for value in record.values())
            assert record["national_id"].isdigit()
            assert record["mobile"].startswith("09")
        assert all(line.startswith("\"") for line in file.read_text("utf-8").splitlines()[1:])

    expected_order = sorted(ordered_records)
    assert ordered_records == expected_order
    assert ordered_records[0] == "0000000001"

    manifest_path = tmp_path / "export_manifest.json"
    assert manifest_path.exists()
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert payload["format"] == "csv"
    assert payload["excel_safety"]["always_quote"] is True
    assert payload["excel_safety"]["formula_guard"] is True
    assert payload["files"][0]["name"].endswith("001.csv")
    assert payload["files"][1]["name"].endswith("002.csv")
    assert manifest.total_rows == 3
    assert manifest.files[0].row_count == 2
    assert manifest.files[1].row_count == 1


def test_chunking_and_naming_determinism(tmp_path: Path) -> None:
    base_time = datetime(2024, 1, 1, 9, 30, tzinfo=timezone.utc)
    rows = [make_row(idx=i, school_code=123000 + i) for i in range(1, 5)]

    csv_exporter = build_exporter(tmp_path / "csv", rows)
    csv_manifest = csv_exporter.run(
        filters=ExportFilters(year=1402, center=1),
        options=ExportOptions(chunk_size=2, output_format="csv"),
        snapshot=ExportSnapshot(marker="snap", created_at=base_time),
        clock_now=base_time,
    )

    xlsx_exporter = build_exporter(tmp_path / "xlsx", rows)
    xlsx_manifest = xlsx_exporter.run(
        filters=ExportFilters(year=1402, center=1),
        options=ExportOptions(chunk_size=2, output_format="xlsx"),
        snapshot=ExportSnapshot(marker="snap", created_at=base_time),
        clock_now=base_time,
    )

    csv_names = [file.name for file in csv_manifest.files]
    assert csv_names == csv_manifest.metadata["files_order"]
    assert csv_names == [
        "export_SABT_V1_1402-1_20240101093000_001.csv",
        "export_SABT_V1_1402-1_20240101093000_002.csv",
    ]
    assert xlsx_manifest.files[0].name == "export_SABT_V1_1402-1_20240101093000_001.xlsx"
    assert xlsx_manifest.files[0].sheets == (("Sheet_001", 2), ("Sheet_002", 2))

    csv_rows = []
    for file in sorted((tmp_path / "csv").glob("*.csv")):
        csv_rows.extend(row["national_id"] for row in _read_csv(file))

    from openpyxl import load_workbook

    workbook = load_workbook(tmp_path / "xlsx" / xlsx_manifest.files[0].name, read_only=True)
    try:
        xlsx_rows: list[str] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                xlsx_rows.append(str(row[0]))
    finally:
        workbook.close()

    assert csv_rows == xlsx_rows
