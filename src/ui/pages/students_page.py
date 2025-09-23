from __future__ import annotations

import asyncio
from typing import Dict, List, Optional
from datetime import datetime

from PyQt5.QtCore import QAbstractTableModel, QModelIndex, Qt, QTimer, QVariant, pyqtSignal
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import (
    QAction,
    QComboBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMenu,
    QPushButton,
    QTableView,
    QToolBar,
    QFileDialog,
    QVBoxLayout,
    QWidget,
)
from qasync import asyncSlot

from src.api.client import APIClient
from src.api.models import StudentDTO
from src.ui.core.event_bus import EventBus
from src.ui.pages.dialogs.confirm_dialog import ConfirmDialog
from src.ui.pages.dialogs.student_dialog import StudentDialog
from src.ui.pages.students_presenter import StudentsPresenter
from src.services.excel_service import ExcelExportService
from src.services.excel_import_service import ExcelImportService, ImportValidationResult
from src.ui.pages.dialogs.export_progress_dialog import ExportProgressDialog, ImportProgressDialog
from src.ui.pages.dialogs.import_preview_dialog import ImportPreviewDialog


class StudentsTableModel(QAbstractTableModel):
    """مدل جدول دانش‌آموزان با سربرگ‌های فارسی."""

    HEADERS = [
        "انتخاب",
        "ردیف",
        "کد شمارنده",
        "نام",
        "نام خانوادگی",
        "کدملی",
        "تلفن",
        "جنسیت",
        "مرکز",
        "وضعیت تحصیل",
        "مقطع",
        "عملیات",
    ]

    def __init__(self) -> None:
        super().__init__()
        self.students: List[StudentDTO] = []
        self.show_checkboxes: bool = False
        self.selected_ids: set[int] = set()

    async def load_data(self, students: List[StudentDTO]) -> None:
        self.beginResetModel()
        self.students = students
        self.endResetModel()

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:  # noqa: N802
        return 0 if parent.isValid() else len(self.students)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:  # noqa: N802
        return 0 if parent.isValid() else len(self.HEADERS)

    def headerData(self, section: int, orientation: int, role: int = Qt.DisplayRole):  # noqa: N802
        if role == Qt.DisplayRole and orientation == Qt.Horizontal:
            return self.HEADERS[section]
        return QVariant()

    def data(self, index: QModelIndex, role: int = Qt.DisplayRole):  # noqa: N802
        if not index.isValid():
            return QVariant()
        s = self.students[index.row()]
        c = index.column()
        if role == Qt.CheckStateRole and c == 0 and self.show_checkboxes:
            return Qt.Checked if s.student_id in self.selected_ids else Qt.Unchecked
        if role != Qt.DisplayRole:
            return QVariant()
        if c == 0:
            return QVariant()
        if c == 1:
            return index.row() + 1
        if c == 2:
            return s.counter
        if c == 3:
            return s.first_name
        if c == 4:
            return s.last_name
        if c == 5:
            return s.national_code
        if c == 6:
            return self._format_phone(s.phone)
        if c == 7:
            return "زن" if s.gender == 0 else "مرد"
        if c == 8:
            centers = {1: "مرکز", 2: "گلستان", 3: "صدرا"}
            return centers.get(int(s.center), "نامشخص")
        if c == 9:
            return "درحال تحصیل" if s.education_status == 1 else "فارغ‌التحصیل"
        if c == 10:
            return s.grade_level
        if c == 11:
            return "✏️ | 🗑️"
        return QVariant()

    def flags(self, index: QModelIndex):  # noqa: D401
        base = super().flags(index)
        if index.isValid() and index.column() == 0 and self.show_checkboxes:
            return base | Qt.ItemIsUserCheckable | Qt.ItemIsEnabled
        return base

    def setData(self, index: QModelIndex, value, role: int = Qt.EditRole):  # noqa: D401, ANN001
        if index.isValid() and index.column() == 0 and role == Qt.CheckStateRole:
            s = self.students[index.row()]
            if value == Qt.Checked:
                self.selected_ids.add(s.student_id)
            else:
                self.selected_ids.discard(s.student_id)
            self.dataChanged.emit(index, index, [Qt.CheckStateRole])
            return True
        return False

    def toggle_selection_mode(self, enabled: bool) -> None:
        self.show_checkboxes = enabled
        if not enabled:
            self.selected_ids.clear()
        self.layoutChanged.emit()

    def get_selected_students(self) -> List[StudentDTO]:
        ids = self.selected_ids
        return [s for s in self.students if s.student_id in ids]

    def select_all(self, select: bool = True) -> None:
        if select:
            self.selected_ids = {s.student_id for s in self.students}
        else:
            self.selected_ids.clear()
        self.layoutChanged.emit()

    @staticmethod
    def _format_phone(phone: str) -> str:
        if phone.startswith("+98"):
            rest = phone[3:]
            return ("0" + rest) if rest else phone
        return phone


class FilterPanel(QWidget):
    """پنل فیلتر پیشرفته با جستجوی لحظه‌ای."""

    filters_changed = pyqtSignal(dict)
    search_requested = pyqtSignal(str)

    def __init__(self) -> None:
        super().__init__()
        self._search_timer = QTimer(self)
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(400)
        self._setup_ui()
        self._setup_connections()

    def _setup_ui(self) -> None:
        layout = QVBoxLayout(self)

        # ردیف جستجو
        search_row = QHBoxLayout()
        self.first_name_input = QLineEdit()
        self.first_name_input.setPlaceholderText("نام...")
        self.last_name_input = QLineEdit()
        self.last_name_input.setPlaceholderText("نام خانوادگی...")
        self.national_code_input = QLineEdit()
        self.national_code_input.setPlaceholderText("کدملی...")
        self.phone_input = QLineEdit()
        self.phone_input.setPlaceholderText("تلفن...")
        self.counter_input = QLineEdit()
        self.counter_input.setPlaceholderText("کد شمارنده...")
        self.search_btn = QPushButton("🔍 جستجو")
        self.refresh_btn = QPushButton("🔄 بروزرسانی")
        search_row.addWidget(QLabel("نام:"))
        search_row.addWidget(self.first_name_input)
        search_row.addWidget(QLabel("نام خانوادگی:"))
        search_row.addWidget(self.last_name_input)
        search_row.addWidget(QLabel("کدملی:"))
        search_row.addWidget(self.national_code_input)
        search_row.addWidget(QLabel("تلفن:"))
        search_row.addWidget(self.phone_input)
        search_row.addWidget(QLabel("کد شمارنده:"))
        search_row.addWidget(self.counter_input)
        search_row.addWidget(self.search_btn)
        search_row.addWidget(self.refresh_btn)
        search_row.addStretch()

        # ردیف فیلترها
        filter_row = QHBoxLayout()
        self.gender_combo = QComboBox()
        self.gender_combo.addItems(["همه", "زن", "مرد"])
        self.center_combo = QComboBox()
        self.center_combo.addItems(["همه", "مرکز", "گلستان", "صدرا"])
        self.status_combo = QComboBox()
        self.status_combo.addItems(["همه", "درحال تحصیل", "فارغ‌التحصیل"])
        self.add_btn = QPushButton("➕ افزودن دانش‌آموز")
        self.add_btn.setStyleSheet("background-color: #27ae60; font-weight: bold;")
        filter_row.addWidget(QLabel("جنسیت:"))
        filter_row.addWidget(self.gender_combo)
        filter_row.addWidget(QLabel("مرکز:"))
        filter_row.addWidget(self.center_combo)
        filter_row.addWidget(QLabel("وضعیت:"))
        filter_row.addWidget(self.status_combo)
        filter_row.addStretch()
        filter_row.addWidget(self.add_btn)

        layout.addLayout(search_row)
        layout.addLayout(filter_row)

    def _setup_connections(self) -> None:
        self.first_name_input.textChanged.connect(self._on_search_changed)
        self.last_name_input.textChanged.connect(self._on_search_changed)
        self.national_code_input.textChanged.connect(self._on_search_changed)
        self.phone_input.textChanged.connect(self._on_search_changed)
        self.counter_input.textChanged.connect(self._on_search_changed)
        # explicit buttons
        self.search_btn.clicked.connect(lambda: self.search_requested.emit(""))
        self.refresh_btn.clicked.connect(lambda: self.search_requested.emit(""))
        # filter combos emit filters_changed
        self.gender_combo.currentIndexChanged.connect(lambda: self.filters_changed.emit(self.get_filters()))
        self.center_combo.currentIndexChanged.connect(lambda: self.filters_changed.emit(self.get_filters()))
        self.status_combo.currentIndexChanged.connect(lambda: self.filters_changed.emit(self.get_filters()))

    def _on_search_changed(self, _):  # noqa: ANN001
        self._search_timer.stop()
        self._search_timer.timeout.connect(lambda: None)
        self._search_timer.timeout.disconnect()
        self._search_timer.timeout.connect(self._emit_search)
        self._search_timer.start()

    def _emit_search(self) -> None:
        self.search_requested.emit(self.name_input.text().strip())

    def get_filters(self) -> Dict:
        filters: Dict = {}
        if self.gender_combo.currentIndex() > 0:
            filters["gender"] = self.gender_combo.currentIndex() - 1
        if self.center_combo.currentIndex() > 0:
            filters["center"] = self.center_combo.currentIndex()
        if self.status_combo.currentIndex() > 0:
            filters["education_status"] = self.status_combo.currentIndex() - 1
        if self.first_name_input.text().strip():
            filters["first_name_search"] = self.first_name_input.text().strip()
        if self.last_name_input.text().strip():
            filters["last_name_search"] = self.last_name_input.text().strip()
        if self.national_code_input.text().strip():
            filters["national_code"] = self.national_code_input.text().strip()
        if self.phone_input.text().strip():
            filters["phone"] = self.phone_input.text().strip()
        if self.counter_input.text().strip():
            filters["counter_search"] = self.counter_input.text().strip()
        return filters


class PaginationWidget(QWidget):
    """ویجت صفحه‌بندی فارسی با ناوبری صفحات."""

    def __init__(self) -> None:
        super().__init__()
        self.current_page = 1
        self.total_pages = 1
        self.total_count = 0
        self.page_size = 20
        self._setup_ui()

    def _setup_ui(self) -> None:
        layout = QHBoxLayout(self)
        self.info_label = QLabel("نمایش 1-20 از 0")
        self.prev_btn = QPushButton("◀ قبلی")
        self.page_label = QLabel("صفحه 1 از 1")
        self.next_btn = QPushButton("بعدی ▶")
        self.page_size_combo = QComboBox()
        self.page_size_combo.addItems(["10 در صفحه", "20 در صفحه", "50 در صفحه", "100 در صفحه"])
        self.page_size_combo.setCurrentText("20 در صفحه")
        layout.addWidget(self.info_label)
        layout.addStretch()
        layout.addWidget(self.prev_btn)
        layout.addWidget(self.page_label)
        layout.addWidget(self.next_btn)
        layout.addStretch()
        layout.addWidget(self.page_size_combo)

    def update_pagination(self, current_page: int, total_count: int, page_size: int) -> None:
        self.current_page = current_page
        self.total_count = total_count
        self.page_size = page_size
        self.total_pages = max(1, (total_count + page_size - 1) // page_size)
        start = (current_page - 1) * page_size + 1 if total_count else 0
        end = min(current_page * page_size, total_count)
        self.info_label.setText(f"نمایش {start}-{end} از {total_count}")
        self.page_label.setText(f"صفحه {current_page} از {self.total_pages}")
        self.prev_btn.setEnabled(current_page > 1)
        self.next_btn.setEnabled(current_page < self.total_pages)


class StudentsPage(QWidget):
    """صفحه کامل مدیریت دانش‌آموزان با جدول، فیلتر و CRUD."""

    def __init__(self, api_client: APIClient, event_bus: EventBus) -> None:
        super().__init__()
        self.api_client = api_client
        self.event_bus = event_bus
        self.presenter = StudentsPresenter(api_client, event_bus)

        # State
        self.current_page = 1
        self.page_size = 20
        self.total_count = 0
        self.filters: Dict = {}

        self._setup_ui()
        self._setup_connections()
        asyncio.create_task(self._load_page())

    def _setup_ui(self) -> None:
        layout = QVBoxLayout(self)

        # Export/Import toolbar
        self.toolbar = QToolBar("دانش‌آموزان", self)
        self._init_toolbar()
        layout.addWidget(self.toolbar)

        # Filters
        self.filter_panel = FilterPanel()
        layout.addWidget(self.filter_panel)

        # Table
        self.table = QTableView(self)
        self.table.setSelectionBehavior(QTableView.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.model = StudentsTableModel()
        self.table.setModel(self.model)
        layout.addWidget(self.table)

        # Pagination
        self.pagination = PaginationWidget()
        layout.addWidget(self.pagination)

    def _setup_connections(self) -> None:
        self.filter_panel.search_btn.clicked.connect(self.on_search_clicked)
        self.filter_panel.refresh_btn.clicked.connect(self.on_refresh_clicked)
        self.filter_panel.add_btn.clicked.connect(self.on_add_clicked)
        self.filter_panel.filters_changed.connect(self.on_filters_changed)
        self.filter_panel.search_requested.connect(self.on_search_triggered)

        self.pagination.prev_btn.clicked.connect(self.on_prev_page)
        self.pagination.next_btn.clicked.connect(self.on_next_page)
        self.pagination.page_size_combo.currentTextChanged.connect(self.on_page_size_changed)

        # Double-click edit as a quick action
        self.table.doubleClicked.connect(self._on_table_double_clicked)
        self.model.dataChanged.connect(lambda *_: self._update_selected_count())

        # Toolbar actions
        self.action_export_current.triggered.connect(lambda: asyncio.create_task(self.export_current_view()))
        self.action_export_all.triggered.connect(lambda: asyncio.create_task(self.export_all()))
        self.action_download_template.triggered.connect(lambda: asyncio.create_task(self.download_template()))
        self.action_import.triggered.connect(lambda: asyncio.create_task(self.import_from_excel()))
        self.action_selection_mode.toggled.connect(self.on_toggle_selection_mode)
        self.action_select_all.triggered.connect(lambda: self.on_select_all(True))
        self.action_select_none.triggered.connect(lambda: self.on_select_all(False))
        self.action_bulk_delete.triggered.connect(lambda: asyncio.create_task(self.bulk_delete()))
        self.action_bulk_export.triggered.connect(lambda: asyncio.create_task(self.export_selected()))

    async def _load_page(self) -> None:
        filt = self.filter_panel.get_filters()
        result, total = await self.presenter.load_students(
            page=self.current_page,
            page_size=self.page_size,
            filters=filt,
            search="",
        )
        await self.model.load_data(result)
        self.total_count = total
        self.pagination.update_pagination(self.current_page, self.total_count, self.page_size)
        self._inject_row_actions()
        self._update_selected_count()

    def _inject_row_actions(self) -> None:
        # ساخت دکمه‌های عملیات در ستون آخر
        for row in range(self.model.rowCount()):
            idx = self.model.index(row, 11)
            w = QWidget(self.table)
            lay = QHBoxLayout(w)
            lay.setContentsMargins(0, 0, 0, 0)
            edit_btn = QPushButton("✏️")
            del_btn = QPushButton("🗑️")
            edit_btn.setFixedWidth(36)
            del_btn.setFixedWidth(36)
            lay.addStretch()
            lay.addWidget(edit_btn)
            lay.addWidget(del_btn)
            lay.addStretch()
            self.table.setIndexWidget(idx, w)

            s = self.model.students[row]
            edit_btn.clicked.connect(lambda _, sid=s.student_id: asyncio.create_task(self._edit_student(sid)))
            del_btn.clicked.connect(lambda _, sid=s.student_id: asyncio.create_task(self._delete_student(sid)))

    # ---------------- actions ----------------
    @asyncSlot()
    async def on_search_clicked(self) -> None:
        self.current_page = 1
        await self._load_page()

    @asyncSlot()
    async def on_refresh_clicked(self) -> None:
        await self._load_page()

    @asyncSlot()
    async def on_add_clicked(self) -> None:
        dlg = StudentDialog(parent=self)
        if dlg.exec_() == dlg.Accepted:
            data = dlg.get_student_data()
            ok = await self.presenter.add_student(data)
            if ok:
                await self._load_page()

    @asyncSlot()
    async def on_filters_changed(self, _=None) -> None:  # noqa: ANN001
        self.current_page = 1
        await self._load_page()

    @asyncSlot()
    async def on_search_triggered(self, _=None) -> None:  # noqa: ANN001
        self.current_page = 1
        await self._load_page()

    @asyncSlot()
    async def on_prev_page(self) -> None:
        if self.current_page > 1:
            self.current_page -= 1
            await self._load_page()

    @asyncSlot()
    async def on_next_page(self) -> None:
        total_pages = max(1, (self.total_count + self.page_size - 1) // self.page_size)
        if self.current_page < total_pages:
            self.current_page += 1
            await self._load_page()

    @asyncSlot()
    async def on_page_size_changed(self, text: str) -> None:
        try:
            self.page_size = int(text.split()[0])
        except Exception:  # noqa: BLE001
            self.page_size = 20
        self.current_page = 1
        await self._load_page()

    def _on_table_double_clicked(self, index: QModelIndex) -> None:
        if not index.isValid():
            return
        # ویرایش با دوبل کلیک
        row = index.row()
        s = self.model.students[row]
        asyncio.create_task(self._edit_student(s.student_id))

    async def _edit_student(self, student_id: int) -> None:
        student: Optional[StudentDTO] = next((x for x in self.model.students if x.student_id == student_id), None)
        if not student:
            return

        dlg = StudentDialog(student=student, parent=self)
        if dlg.exec_() != dlg.Accepted:
            return

        data = dlg.get_student_data()
        ok = await self.presenter.update_student(student_id, data)
        if ok:
            await self._load_page()


    async def _delete_student(self, student_id: int) -> None:
        dlg = ConfirmDialog("آیا از حذف این دانش‌آموز مطمئن هستید؟", parent=self)
        if dlg.exec_() == dlg.Accepted:
            ok = await self.presenter.delete_student(student_id)
            if ok:
                await self._load_page()

    @asyncSlot()
    async def bulk_delete(self) -> None:
        students = self.model.get_selected_students()
        if not students:
            return
        dlg = ConfirmDialog(f"حذف {len(students)} دانش‌آموز انتخاب شده؟", parent=self)
        if dlg.exec_() != dlg.Accepted:
            return
        for s in students:
            await self.presenter.delete_student(s.student_id)
        await self._load_page()

    # ---------------- toolbar helpers ----------------
    def _init_toolbar(self) -> None:
        # Export dropdown
        export_menu = QMenu(self)
        self.action_export_current = QAction("صادرات نمای فعلی", self)
        self.action_export_all = QAction("صادرات همه دانش‌آموزان", self)
        self.action_download_template = QAction("دانلود قالب ورود", self)
        export_menu.addAction(self.action_export_current)
        export_menu.addAction(self.action_export_all)
        export_menu.addSeparator()
        export_menu.addAction(self.action_download_template)

        self.export_button = QPushButton("صادرات")
        self.export_button.setMenu(export_menu)
        self.toolbar.addWidget(self.export_button)

        # Import button
        self.action_import = QAction("ورود از Excel", self)
        self.toolbar.addAction(self.action_import)

        self.toolbar.addSeparator()

        # Selection mode toggle and bulk actions
        self.action_selection_mode = QAction("حالت انتخاب گروهی", self)
        self.action_selection_mode.setCheckable(True)
        self.toolbar.addAction(self.action_selection_mode)

        self.action_select_all = QAction("انتخاب همه", self)
        self.action_select_none = QAction("لغو انتخاب", self)
        self.action_bulk_delete = QAction("حذف انتخاب شده‌ها", self)
        self.action_bulk_export = QAction("صادرات انتخاب شده‌ها", self)
        for act in [self.action_select_all, self.action_select_none, self.action_bulk_delete, self.action_bulk_export]:
            act.setVisible(False)
            self.toolbar.addAction(act)

        self.selected_count_label = QLabel("0 مورد انتخاب شده")
        self.selected_count_label.setVisible(False)
        self.toolbar.addWidget(self.selected_count_label)

    def on_toggle_selection_mode(self, enabled: bool) -> None:
        self.model.toggle_selection_mode(enabled)
        for act in [self.action_select_all, self.action_select_none, self.action_bulk_delete, self.action_bulk_export]:
            act.setVisible(enabled)
        self.selected_count_label.setVisible(enabled)
        self._update_selected_count()

    def on_select_all(self, select: bool) -> None:
        self.model.select_all(select)
        self._update_selected_count()

    def _update_selected_count(self) -> None:
        self.selected_count_label.setText(f"{len(self.model.selected_ids)} مورد انتخاب شده")

    # ---------------- export/import actions ----------------
    async def export_current_view(self) -> None:
        students = self.model.students
        await self._export_students(students, prefix="دانش_آموزان_فیلتر_شده")

    async def export_all(self) -> None:
        # Try paginated fetch of all pages (simple: get all without filters)
        result = await self.presenter.load_students(page=1, page_size=10_000, filters={}, search="")
        students = result[0]
        await self._export_students(students, prefix="دانش_آموزان")

    async def export_selected(self) -> None:
        students = self.model.get_selected_students()
        if not students:
            return
        await self._export_students(students, prefix="دانش_آموزان_انتخاب_شده")

    async def _export_students(self, students: List[StudentDTO], prefix: str) -> None:
        if not students:
            return
        dt = datetime.now().strftime("%Y_%m_%d_%H_%M")
        default = f"{prefix}_{dt}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "ذخیره فایل اکسل", default, "Excel Files (*.xlsx)")
        if not path:
            return
        svc = ExcelExportService()
        dlg = ExportProgressDialog(total_rows=len(students), parent=self)
        dlg.show()

        def progress(done: int, total: int) -> bool:
            dlg.update_progress(done, total)
            return not dlg.is_cancelled()

        try:
            await svc.export_students(students, path, progress_callback=progress)
        except Exception:
            pass
        dlg.close()

    async def download_template(self) -> None:
        # Build a minimal template using the export headers subset useful for import
        import openpyxl
        from openpyxl.worksheet.datavalidation import DataValidation

        headers = [
            "نام",
            "نام خانوادگی",
            "کدملی",
            "تلفن",
            "تاریخ تولد",
            "جنسیت",
            "وضعیت تحصیل",
            "نوع ثبت‌نام",
            "مرکز",
            "مقطع تحصیلی",
            "نوع مدرسه",
            "کد مدرسه",
        ]
        path, _ = QFileDialog.getSaveFileName(self, "ذخیره قالب ورود", "قالب_ورود_دانش_آموزان.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "قالب"
        ws.sheet_view.rightToLeft = True
        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h)
        # sample
        sample = ["علی", "احمدی", "0012345678", "09123456789", "2005/01/01", "مرد", "در حال تحصیل", "عادی", "مرکز", "konkoori", "عادی", ""]
        for c, v in enumerate(sample, 1):
            ws.cell(row=2, column=c, value=v)
        # drop-down validations
        dv_gender = DataValidation(type="list", formula1='"زن,مرد"', allow_blank=False)
        dv_edu = DataValidation(type="list", formula1='"در حال تحصیل,فارغ‌التحصیل"', allow_blank=False)
        dv_reg = DataValidation(type="list", formula1='"عادی,شهید,حکمت"', allow_blank=True)
        dv_center = DataValidation(type="list", formula1='"مرکز,گلستان,صدرا"', allow_blank=True)
        dv_school_type = DataValidation(type="list", formula1='"عادی,مدرسه‌ای"', allow_blank=True)
        ws.add_data_validation(dv_gender)
        ws.add_data_validation(dv_edu)
        ws.add_data_validation(dv_reg)
        ws.add_data_validation(dv_center)
        ws.add_data_validation(dv_school_type)
        dv_gender.add(ws["F2:F10000"])  # جنسیت
        dv_edu.add(ws["G2:G10000"])    # وضعیت تحصیل
        dv_reg.add(ws["H2:H10000"])    # نوع ثبت نام
        dv_center.add(ws["I2:I10000"]) # مرکز
        dv_school_type.add(ws["K2:K10000"])  # نوع مدرسه
        wb.save(path)

    async def import_from_excel(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "انتخاب فایل اکسل", "", "Excel Files (*.xlsx)")
        if not path:
            return
        validator = ExcelImportService()
        dlg = ImportProgressDialog("اعتبارسنجی فایل", total_rows=0, parent=self)
        dlg.show()

        def progress(done: int, total: int) -> bool:
            dlg.update_progress(done, total, phase="اعتبارسنجی")
            return not dlg.is_cancelled()

        result: ImportValidationResult = await validator.validate_import_file(path, progress_callback=progress)
        dlg.close()
        if not result.success:
            # Could show a message box; for now ignore
            return

        preview = ImportPreviewDialog(result, parent=self)
        if preview.exec_() != preview.Accepted:
            return

        # If user confirmed import
        # Button signal mapping: accept on import
        # We assume caller connected .import_btn externally; for brevity we check property
        # Batch import
        import_progress = ImportProgressDialog("ورود اطلاعات", total_rows=len(result.valid_rows), parent=self)
        import_progress.show()

        imported = 0
        for i, row in enumerate(result.valid_rows, start=1):
            if import_progress.is_cancelled():
                break
            data = row["data"]
            ok = await self.presenter.add_student(data)
            imported += 1 if ok else 0
            import_progress.update_progress(i, len(result.valid_rows), phase="ورود")
        import_progress.close()
        await self._load_page()
