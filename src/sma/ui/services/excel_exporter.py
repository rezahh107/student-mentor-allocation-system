from __future__ import annotations

from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ...core.models import Mentor, Student


class ExcelExporter:
    """Utility responsible for exporting allocation results to Excel."""

    def __init__(self) -> None:
        self.workbook: Workbook | None = None

    async def export_allocation_results(
        self,
        results: Dict,
        students: List[Student],
        mentors: List[Mentor],
        file_path: str,
    ) -> bool:
        try:
            workbook = Workbook()
            self.workbook = workbook
            workbook.remove(workbook.active)

            self._create_summary_sheet(workbook, results)

            assignments = results.get("assignments", [])
            if assignments:
                self._create_assignments_sheet(workbook, assignments, students, mentors)

            errors = results.get("errors", [])
            if errors:
                self._create_errors_sheet(workbook, errors, students)

            workbook.save(file_path)
            return True
        except Exception as exc:  # noqa: BLE001
            print(f"خطا در صادرات Excel: {exc}")
            return False

    # ------------------------------------------------------------------
    # Sheet builders
    # ------------------------------------------------------------------
    def _create_summary_sheet(self, wb: Workbook, results: Dict) -> None:
        ws = wb.create_sheet("خلاصه", 0)
        ws["A1"] = "گزارش تخصیص دانش‌آموزان"
        ws["A1"].font = Font(size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:D1")

        total = results.get("successful", 0) + results.get("failed", 0)
        success = results.get("successful", 0)
        failed = results.get("failed", 0)
        rate = f"{(success / total * 100):.1f}%" if total else "0%"

        summary_data = [
            ["تعداد کل دانش‌آموزان", total],
            ["تخصیص موفق", success],
            ["تخصیص ناموفق", failed],
            ["درصد موفقیت", rate],
        ]

        for row_index, (label, value) in enumerate(summary_data, start=3):
            ws.cell(row=row_index, column=1, value=label)
            ws.cell(row=row_index, column=2, value=value)

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def _create_assignments_sheet(
        self,
        wb: Workbook,
        assignments: List[Dict],
        students: List[Student],
        mentors: List[Mentor],
    ) -> None:
        ws = wb.create_sheet("تخصیصات موفق")
        headers = [
            "ردیف",
            "نام دانش‌آموز",
            "مقطع",
            "جنسیت",
            "نام پشتیبان",
            "امتیاز اولویت",
        ]
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        students_map = {student.id: student for student in students}
        mentors_map = {mentor.id: mentor for mentor in mentors}

        for row, assignment in enumerate(assignments, start=2):
            student = students_map.get(assignment["student_id"])
            mentor = mentors_map.get(assignment["mentor_id"])
            if not student or not mentor:
                continue

            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=student.name)
            ws.cell(row=row, column=3, value=f"پایه {student.grade_level}")
            ws.cell(row=row, column=4, value="دختر" if student.gender == 0 else "پسر")
            ws.cell(row=row, column=5, value=mentor.name)
            ws.cell(row=row, column=6, value=assignment.get("priority_score", 0))

        for idx in range(1, len(headers) + 1):
            column_letter = chr(64 + idx)
            ws.column_dimensions[column_letter].width = 20

    def _create_errors_sheet(
        self,
        wb: Workbook,
        errors: List[Dict],
        students: List[Student],
    ) -> None:
        ws = wb.create_sheet("خطاها")
        headers = ["ردیف", "شناسه دانش‌آموز", "نام", "علت"]
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        students_map = {student.id: student for student in students}
        for row, error in enumerate(errors, start=2):
            student = students_map.get(error.get("student_id"))
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=error.get("student_id"))
            ws.cell(row=row, column=3, value=getattr(student, "name", "نامشخص"))
            ws.cell(row=row, column=4, value=error.get("reason", "نامشخص"))

        for idx in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + idx)].width = 25
