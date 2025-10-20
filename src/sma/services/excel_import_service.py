from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from sma.api.models import validate_iranian_phone, validate_national_code


@dataclass
class StudentRowValidationResult:
    is_valid: bool
    student_data: Dict[str, Any] = field(default_factory=dict)
    errors: List[str] = field(default_factory=list)


@dataclass
class ImportValidationResult:
    success: bool
    total_rows: int = 0
    valid_rows: List[Dict[str, Any]] = field(default_factory=list)  # {row_number, data}
    invalid_rows: List[Dict[str, Any]] = field(default_factory=list)  # {row_number, data, errors}
    error: Optional[str] = None


PERSIAN_HEADERS = [
    "نام",
    "نام خانوادگی",
    "کدملی",
    "تلفن",
    "تاریخ تولد",
    "جنسیت",
    "وضعیت تحصیل",
    "نوع ثبت‌نام",
    "مرکز",
    "مقطع تحصیلی",
    "نوع مدرسه",
    "کد مدرسه",
]


class ExcelImportService:
    """سرویس اعتبارسنجی و ورود اکسل برای دانش‌آموزان."""

    def __init__(self) -> None:
        self.required_headers = [
            "نام",
            "نام خانوادگی",
            "کدملی",
            "تلفن",
            "تاریخ تولد",
            "جنسیت",
            "وضعیت تحصیل",
        ]

    def check_required_headers(self, headers: List[str]) -> List[str]:
        return [h for h in self.required_headers if h not in headers]

    async def validate_import_file(
        self, filepath: str, progress_callback: Optional[callable] = None
    ) -> ImportValidationResult:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        missing = self.check_required_headers(headers)
        if missing:
            return ImportValidationResult(
                success=False, error=f"ستون‌های مورد نیاز یافت نشد: {', '.join(missing)}"
            )

        valid_rows: List[Dict[str, Any]] = []
        invalid_rows: List[Dict[str, Any]] = []
        seen_national: set[str] = set()
        total = max(0, ws.max_row - 1)

        for r in range(2, ws.max_row + 1):
            if progress_callback:
                cont = progress_callback(r - 1, total)
                if hasattr(cont, "__await__"):
                    cont = await cont
                if cont is False:
                    break

            row_vals = [cell.value for cell in ws[r]]
            result = self.validate_student_row(row_vals, headers)
            if result.is_valid:
                nat = str(result.student_data.get("national_code", ""))
                if nat in seen_national:
                    invalid_rows.append({
                        "row_number": r,
                        "data": row_vals,
                        "errors": ["کدملی تکراری در فایل"],
                    })
                else:
                    valid_rows.append({"row_number": r, "data": result.student_data})
                    seen_national.add(nat)
            else:
                invalid_rows.append({"row_number": r, "data": row_vals, "errors": result.errors})

        return ImportValidationResult(
            success=True,
            total_rows=total,
            valid_rows=valid_rows,
            invalid_rows=invalid_rows,
        )

    def validate_student_row(self, row: List[Any], headers: List[str]) -> StudentRowValidationResult:
        data: Dict[str, Any] = {}
        errors: List[str] = []

        map_index = {h: i for i, h in enumerate(headers)}

        def get(h: str) -> Any:
            idx = map_index.get(h)
            return row[idx] if idx is not None and idx < len(row) else None

        # Required: first/last/national_code/phone/birth_date/gender/education_status
        first = (get("نام") or "").strip() if isinstance(get("نام"), str) else get("نام") or ""
        last = (get("نام خانوادگی") or "").strip() if isinstance(get("نام خانوادگی"), str) else get("نام خانوادگی") or ""
        national = str(get("کدملی") or "").strip()
        phone = str(get("تلفن") or "").strip()
        birth = get("تاریخ تولد")
        gender_txt = str(get("جنسیت") or "").strip()
        edu_txt = str(get("وضعیت تحصیل") or "").strip()

        if not first:
            errors.append("نام الزامی است")
        if not last:
            errors.append("نام خانوادگی الزامی است")
        if not validate_national_code(national):
            errors.append("کدملی نامعتبر است")
        if not validate_iranian_phone(phone):
            errors.append("تلفن نامعتبر است")

        # Parse date (supports Excel date, datetime, or string YYYY/MM/DD)
        parsed_birth: Optional[date] = None
        if birth is None:
            errors.append("تاریخ تولد الزامی است")
        else:
            try:
                if isinstance(birth, datetime):
                    parsed_birth = birth.date()
                elif isinstance(birth, date):
                    parsed_birth = birth
                else:
                    # try split
                    s = str(birth).replace("-", "/")
                    parts = [int(p) for p in s.split("/")]
                    if len(parts) == 3:
                        y, m, d = parts
                        parsed_birth = date(y, m, d)
            except Exception:
                errors.append("فرمت تاریخ تولد نامعتبر است")

        gender_map = {"زن": 0, "مرد": 1}
        edu_map = {"فارغ‌التحصیل": 0, "در حال تحصیل": 1, "درحال تحصیل": 1}
        reg_map = {"عادی": 0, "شهید": 1, "حکمت": 2}
        center_map = {"مرکز": 1, "گلستان": 2, "صدرا": 3}
        school_type_map = {"عادی": "normal", "مدرسه‌ای": "school", "مدرسه اي": "school"}

        gender = gender_map.get(gender_txt)
        if gender is None:
            errors.append("مقدار جنسیت نامعتبر است (زن/مرد)")

        edu = edu_map.get(edu_txt)
        if edu is None:
            errors.append("وضعیت تحصیل نامعتبر است")

        reg_txt = str(get("نوع ثبت‌نام") or "").strip()
        reg = reg_map.get(reg_txt, 0)

        center_txt = str(get("مرکز") or "").strip()
        center = center_map.get(center_txt, 1)

        grade_level = str(get("مقطع تحصیلی") or "").strip() or "konkoori"
        school_type_txt = str(get("نوع مدرسه") or "").strip() or "عادی"
        school_type = school_type_map.get(school_type_txt, "normal")
        school_code = str(get("کد مدرسه") or "").strip() or None

        if errors:
            return StudentRowValidationResult(is_valid=False, student_data={}, errors=errors)

        data = {
            "first_name": first,
            "last_name": last,
            "national_code": national,
            "phone": phone,
            "birth_date": parsed_birth.isoformat() if parsed_birth else None,
            "gender": gender,
            "education_status": edu,
            "registration_status": reg,
            "center": center,
            "grade_level": grade_level,
            "school_type": school_type,
            "school_code": school_code,
        }
        return StudentRowValidationResult(is_valid=True, student_data=data, errors=[])
