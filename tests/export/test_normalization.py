from __future__ import annotations

import csv
from dataclasses import replace
from datetime import datetime, timezone

from openpyxl import load_workbook

from sma.phase6_import_to_sabt.export_writer import EXPORT_COLUMNS
from sma.phase6_import_to_sabt.models import ExportFilters, ExportOptions, ExportSnapshot
from tests.export.helpers import build_exporter, make_row


def test_nfkc_digit_folding_and_y_k_unification(tmp_path) -> None:
    base_time = datetime(2024, 3, 23, 7, 30, tzinfo=timezone.utc)
    row = replace(
        make_row(idx=11, school_code=7),
        first_name="\u200dكريم",
        last_name="يحيى",
        mentor_name="نام‌\u200c",
        mentor_id="٠٠١٢٣",
        mobile="۰۹۱۲۳۴۵۶۷۸۹",
    )

    exporter = build_exporter(tmp_path, [row])
    manifest_csv = exporter.run(
        filters=ExportFilters(year=1402, center=1),
        options=ExportOptions(output_format="csv"),
        snapshot=ExportSnapshot(marker="snap", created_at=base_time),
        clock_now=base_time,
    )

    csv_path = tmp_path / manifest_csv.files[0].name
    with csv_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        data = next(reader)
    assert data["first_name"] == "کریم", data
    assert data["last_name"].startswith("یح"), data
    assert "ی" in data["last_name"], data
    assert data["mentor_name"] == "نام", data
    assert data["mentor_id"].isdigit(), data
    assert data["mobile"].startswith("09"), data
    assert data["school_code"] == "000007", data

    exporter_xlsx = build_exporter(tmp_path / "xlsx", [row])
    manifest_xlsx = exporter_xlsx.run(
        filters=ExportFilters(year=1402, center=1),
        options=ExportOptions(output_format="xlsx"),
        snapshot=ExportSnapshot(marker="snap", created_at=base_time),
        clock_now=base_time,
    )

    xlsx_path = tmp_path / "xlsx" / manifest_xlsx.files[0].name
    workbook = load_workbook(xlsx_path, read_only=True)
    try:
        sheet = workbook.active
        row_values = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        mapped = dict(zip(EXPORT_COLUMNS, row_values))
        assert str(mapped["first_name"]) == "کریم"
        assert str(mapped["last_name"]).startswith("یح"), mapped
        assert str(mapped["mentor_name"]) == "نام"
        assert str(mapped["mentor_id"]).isdigit()
        assert str(mapped["mobile"]).startswith("09")
        assert str(mapped["school_code"]) == "000007"
    finally:
        workbook.close()
