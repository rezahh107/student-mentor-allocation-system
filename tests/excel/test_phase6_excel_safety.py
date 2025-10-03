from __future__ import annotations

import time
from pathlib import Path
from typing import Callable

import pytest
from openpyxl import load_workbook

from phase6_import_to_sabt.xlsx.writer import EXPORT_COLUMNS, XLSXStreamWriter


def _retry(action: Callable[[], None], *, attempts: int = 3, base_delay: float = 0.001) -> None:
    errors: list[str] = []
    for attempt in range(1, attempts + 1):
        try:
            action()
            return
        except AssertionError as exc:
            errors.append(str(exc))
            if attempt == attempts:
                raise AssertionError("; ".join(errors))
            delay = base_delay * (2 ** (attempt - 1)) + (attempt * 0.0001)
            time.sleep(delay)


@pytest.fixture
def export_path(tmp_path: Path) -> Path:
    target = tmp_path / "test_export.xlsx"
    yield target
    for leftover in tmp_path.iterdir():
        if leftover.is_file():
            leftover.unlink()


def _build_row() -> dict[str, str]:
    long_text = "تست" + "1" * 33000
    row = {column: "" for column in EXPORT_COLUMNS}
    row.update(
        {
            "national_id": "۰۰۱۲۳۴۵۶۷۸",
            "counter": "۱۴۰۳۷۳۰۰۰۱",
            "first_name": "=SUM(A1)\n",
            "last_name": "+Hello",
            "gender": "1",
            "mobile": "۰۹۱۲۳۴۵۶۷۸۹",
            "reg_center": "1",
            "reg_status": "3",
            "group_code": "12\r\n34",
            "student_type": "0",
            "school_code": "123",
            "mentor_id": "A123",
            "mentor_name": long_text,
            "mentor_mobile": "09120000000",
            "allocation_date": "2024-01-01T00:00:00Z",
            "year_code": "1402",
        }
    )
    return row


def test_excel_cells_are_sanitized(export_path: Path) -> None:
    writer = XLSXStreamWriter(chunk_size=10)
    context = {"path": str(export_path)}
    rows = [_build_row()]

    artifact = writer.write(rows, export_path)
    context["artifact"] = {
        "sha256": artifact.sha256,
        "byte_size": artifact.byte_size,
        "row_counts": artifact.row_counts,
    }

    assert export_path.exists(), f"Export file missing: {context}"
    wb = load_workbook(export_path, data_only=False)
    sheet = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in sheet[1]]
    row_map = {headers[idx - 1]: cell for idx, cell in enumerate(sheet[2], start=1)}

    def _assert_no_formulas() -> None:
        for column, cell in row_map.items():
            assert cell.data_type != "f", f"Formula detected in {column}: {cell.value}"

    _retry(_assert_no_formulas)

    first_name = row_map["first_name"].value
    assert isinstance(first_name, str), f"first_name not string: {first_name!r}"
    assert first_name.startswith("'"), f"Formula guard missing: {first_name!r}"
    assert "\n" not in first_name, f"Newlines not collapsed: {first_name!r}"

    mentor_name = row_map["mentor_name"].value
    assert isinstance(mentor_name, str), "mentor_name must be text"
    assert len(mentor_name) <= 32767, f"mentor_name exceeds Excel limit: {len(mentor_name)}"
    assert mentor_name.endswith("…"), "mentor_name should end with ellipsis after truncation"

    national_id = row_map["national_id"].value
    counter_value = row_map["counter"].value
    assert national_id.isdigit(), f"national_id not ASCII digits: {national_id!r}"
    assert counter_value.isdigit(), f"counter not ASCII digits: {counter_value!r}"

    group_code = row_map["group_code"].value
    assert any("\u06f0" <= ch <= "\u06f9" for ch in group_code), (
        "group_code should contain Persian digits",
        group_code,
    )

    leftovers = list(export_path.parent.glob("*.part"))
    assert not leftovers, f"Temporary files leaked: {leftovers}"
