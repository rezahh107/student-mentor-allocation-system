from __future__ import annotations

import asyncio
import os

import pytest

from tests.ui import _headless

_headless.require_ui()

pytestmark = [
    pytest.mark.ui,
    pytest.mark.skipif(
        os.name == "nt", reason="Import/Export UI tests require a GUI environment on Windows"
    ),
]
if _headless.PYTEST_SKIP_MARK is not None:
    pytestmark.append(_headless.PYTEST_SKIP_MARK)

import openpyxl
from PyQt5.QtWidgets import QFileDialog

from src.api.client import APIClient
from src.ui.core.event_bus import EventBus
from src.ui.pages.students_page import StudentsPage


@pytest.mark.asyncio
async def test_download_template_and_import_preview(qtbot, tmp_path, monkeypatch):
    client = APIClient(use_mock=True)
    page = StudentsPage(client, EventBus())
    qtbot.addWidget(page)
    page.show()

    # Download template
    template_path = tmp_path / "template.xlsx"
    monkeypatch.setattr(QFileDialog, "getSaveFileName", lambda *a, **k: (str(template_path), "Excel Files (*.xlsx)"))
    await page.download_template()
    assert template_path.exists()
    wb = openpyxl.load_workbook(str(template_path))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "ظ†ط§ظ…" in headers and "ع©ط¯ظ…ظ„غŒ" in headers

    # Create sample import file with 2 valid rows
    import_path = tmp_path / "import.xlsx"
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.append(headers)
    ws2.append(["ط¹ظ„غŒ", "ط§ط­ظ…ط¯غŒ", "0012345678", "09123456789", "2005/01/01", "ظ…ط±ط¯", "ط¯ط± ط­ط§ظ„ طھط­طµغŒظ„", "ط¹ط§ط¯غŒ", "ظ…ط±ع©ط²", "konkoori", "ط¹ط§ط¯غŒ", ""])
    ws2.append(["ط²ظ‡ط±ط§", "ظ…ط­ظ…ط¯غŒ", "0012345679", "09121234567", "2006/01/01", "ط²ظ†", "ط¯ط± ط­ط§ظ„ طھط­طµغŒظ„", "ط¹ط§ط¯غŒ", "ظ…ط±ع©ط²", "motavassete2", "ط¹ط§ط¯غŒ", ""])
    wb2.save(str(import_path))

    # Patch open dialog to return file path and auto-accept preview dialog
    from src.ui.pages.dialogs import import_preview_dialog as ipd
    monkeypatch.setattr(QFileDialog, "getOpenFileName", lambda *a, **k: (str(import_path), "Excel Files (*.xlsx)"))
    monkeypatch.setattr(ipd.ImportPreviewDialog, "exec_", lambda self: self.Accepted)

    await page.import_from_excel()
    # After import, a refresh happens
    assert page.model.rowCount() > 0

