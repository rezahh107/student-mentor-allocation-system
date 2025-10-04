from __future__ import annotations

from dataclasses import asdict, replace
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from phase6_import_to_sabt.export_writer import ExportWriter
from tests.export.helpers import make_row


def test_default_xlsxwriter_write_only(tmp_path: Path) -> None:
    writer = ExportWriter(sensitive_columns=("national_id", "mentor_id"))
    rows = [
        replace(make_row(idx=1), mentor_name="=SUM(A1)", mentor_id="۰۱۲۳۴۵"),
        replace(make_row(idx=2), mentor_name="@cmd", mentor_id="۰۱۲۳۴۶"),
    ]

    payload = [asdict(row) for row in rows]
    result = writer.write_xlsx(payload, path_factory=lambda index: tmp_path / f"export_{index}.xlsx")
    artifact = result.files[0]
    assert artifact.sheets == (("Sheet_001", len(rows)),)
    assert result.excel_safety["backend"] == "xlsxwriter"

    path = tmp_path / "export_1.xlsx"
    with ZipFile(path, "r") as archive:
        assert "xl/worksheets/sheet1.xml" in archive.namelist(), archive.namelist()

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        sheet = workbook.active
        assert sheet.max_row == len(rows) + 1
        cells = list(sheet.iter_rows(min_row=2, max_row=2))[0]
        national_id_cell = cells[0]
        mentor_name_cell = cells[12]
        assert national_id_cell.number_format == "@"
        assert mentor_name_cell.value.startswith("'")
    finally:
        workbook.close()
