from __future__ import annotations

import csv
from datetime import datetime, timezone

from openpyxl import load_workbook

from sma.phase6_import_to_sabt.models import ExportFilters, ExportOptions, ExportSnapshot
from tests.export.helpers import build_exporter, make_row


def test_logical_parity_between_csv_and_xlsx(tmp_path) -> None:
    base_time = datetime(2024, 3, 24, 6, 0, tzinfo=timezone.utc)
    rows = [make_row(idx=i) for i in range(1, 6)]

    exporter_csv = build_exporter(tmp_path / "csv", rows)
    manifest_csv = exporter_csv.run(
        filters=ExportFilters(year=1402, center=1),
        options=ExportOptions(chunk_size=2, output_format="csv"),
        snapshot=ExportSnapshot(marker="snap", created_at=base_time),
        clock_now=base_time,
    )

    csv_values: list[list[str]] = []
    for manifest_file in manifest_csv.files:
        csv_path = tmp_path / "csv" / manifest_file.name
        with csv_path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            csv_values.extend([[value for value in row.values()] for row in reader])

    exporter_xlsx = build_exporter(tmp_path / "xlsx", rows)
    manifest_xlsx = exporter_xlsx.run(
        filters=ExportFilters(year=1402, center=1),
        options=ExportOptions(chunk_size=2, output_format="xlsx"),
        snapshot=ExportSnapshot(marker="snap", created_at=base_time),
        clock_now=base_time,
    )

    xlsx_values: list[list[str]] = []
    xlsx_path = tmp_path / "xlsx" / manifest_xlsx.files[0].name
    workbook = load_workbook(xlsx_path, read_only=True)
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                xlsx_values.append(["" if value is None else str(value) for value in row])
    finally:
        workbook.close()

    assert csv_values == xlsx_values, {
        "csv": csv_values,
        "xlsx": xlsx_values,
    }
