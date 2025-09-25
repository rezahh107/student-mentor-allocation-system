import json
from pathlib import Path
from typing import Iterable, Tuple

import pytest
from openpyxl import load_workbook

from src.tools.export.xlsx_exporter import XLSXAllocationExporter

_GOLDEN_JSON = Path(__file__).resolve().parents[1] / "golden" / "export" / "xlsx_expected.json"


@pytest.fixture
def sample_rows() -> list[Tuple[object, ...]]:
    return [
        (
            1,
            "AL-001",
            "۱۴۰۲",
            "۱۲۳۴۵",
            77,
            "queued",
            "=SUM(A1:A2)",
            "2024-01-02T03:04:05",
        ),
        (
            2,
            "AL-002",
            "1402",
            "00123",
            0,
            "sent",
            "+HACK",
            "",
        ),
    ]


def _patch_stream(monkeypatch: pytest.MonkeyPatch, rows: Iterable[Tuple[object, ...]]) -> None:
    def fake_stream(self: XLSXAllocationExporter, *, session: object):
        return iter(rows)

    monkeypatch.setattr(XLSXAllocationExporter, "_stream_rows", fake_stream, raising=False)


def test_xlsx_exporter_matches_golden(tmp_path: Path, monkeypatch: pytest.MonkeyPatch, sample_rows: list[Tuple[object, ...]]) -> None:
    _patch_stream(monkeypatch, sample_rows)
    exporter = XLSXAllocationExporter(chunk_size=1, excel_safe=True)
    output = tmp_path / "allocations.xlsx"
    exporter.export(session=object(), output=output)

    with _GOLDEN_JSON.open(encoding="utf-8") as handle:
        expected = json.load(handle)

    workbook = load_workbook(output, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    assert list(rows[0]) == expected["headers"]
    assert [list(row) for row in rows[1:]] == expected["rows"]
