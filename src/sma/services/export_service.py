from __future__ import annotations

from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from sma.services.analytics_service import DashboardData


class DataExportService:
    """قابلیت‌های گسترده خروجی برای داده‌های داشبورد."""

    async def export_to_excel(self, dashboard_data: DashboardData, filename: str, include_charts: bool = True) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "خلاصه آمار"

        # Summary
        ws.append(["شاخص", "مقدار"])
        ws.append(["کل دانش‌آموزان", dashboard_data.total_students])
        ws.append(["ثبت‌نام فعال", dashboard_data.active_students])
        ws.append(["در انتظار تخصیص", dashboard_data.pending_allocations])
        ws.append(["نرخ رشد", dashboard_data.growth_rate])

        # Autofit
        for col in ws.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 2, 50)

        # Detailed
        det = wb.create_sheet("داده‌های تفصیلی")
        det.append(["first_name", "last_name", "gender", "center", "created_at"])
        # Note: Detailed student rows would require access to the list; omitted for brevity.

        if include_charts:
            # Placeholder sheet for charts (not implemented)
            wb.create_sheet("نمودارها")

        wb.save(filename)
        return filename

    async def export_to_csv(self, dashboard_data: DashboardData, filename: str) -> str:
        # ساده‌سازی: CSV خلاصه آمار
        import csv

        with open(filename, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["شاخص", "مقدار"])
            writer.writerow(["کل دانش‌آموزان", dashboard_data.total_students])
            writer.writerow(["ثبت‌نام فعال", dashboard_data.active_students])
            writer.writerow(["در انتظار تخصیص", dashboard_data.pending_allocations])
            writer.writerow(["نرخ رشد", dashboard_data.growth_rate])
        return filename

