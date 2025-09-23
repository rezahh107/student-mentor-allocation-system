from __future__ import annotations

import asyncio
from datetime import datetime
from typing import Callable, Iterable, List, Optional

import jdatetime
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.api.models import StudentDTO


EXCEL_HEADERS = [
    "شناسه",
    "کد شمارنده",
    "نام",
    "نام خانوادگی",
    "کدملی",
    "تلفن",
    "تاریخ تولد",
    "جنسیت",
    "وضعیت تحصیل",
    "نوع ثبت‌نام",
    "مرکز",
    "مقطع تحصیلی",
    "نوع مدرسه",
    "کد مدرسه",
    "تاریخ ثبت‌نام",
    "آخرین بروزرسانی",
    "وضعیت تخصیص",
]


class ExcelExportService:
    """سرویس خروجی اکسل برای لیست دانش‌آموزان با فرمت فارسی و RTL."""

    def __init__(self) -> None:
        self.persian_font = Font(name="B Nazanin", size=11)
        self.header_font = Font(name="B Nazanin", size=12, bold=True)

    async def export_students(
        self,
        students: Iterable[StudentDTO],
        filename: str,
        progress_callback: Optional[Callable[[int, int], "asyncio.Future[bool] | bool"]] = None,
    ) -> str:
        """خروجی اکسل با فرمت‌بندی فارسی و پیشرفت.

        progress_callback: تابعی که (done, total) می‌گیرد و می‌تواند True/False برگرداند
        برای ادامه/لغو.
        """

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "دانش‌آموزان"
        ws.sheet_view.rightToLeft = True

        # Header row
        for col, header in enumerate(EXCEL_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        students_list = list(students)
        total = len(students_list)
        for idx, s in enumerate(students_list, start=2):
            # check cancel
            if progress_callback is not None:
                cont = progress_callback(idx - 1, total)
                if asyncio.iscoroutine(cont):  # type: ignore[attr-defined]
                    cont = await cont  # type: ignore[assignment]
                if cont is False:
                    raise asyncio.CancelledError()

            row = self._format_student_for_excel(s)
            for col, value in enumerate(row, 1):
                ws.cell(row=idx, column=col, value=value)

        # Freeze header
        ws.freeze_panes = "A2"

        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[letter].width = min(max_length + 2, 50)

        # Header filters
        ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_HEADERS))}1"

        wb.save(filename)
        return filename

    def _format_student_for_excel(self, s: StudentDTO) -> List[str]:
        gender_map = {0: "زن", 1: "مرد"}
        edu_map = {0: "فارغ‌التحصیل", 1: "در حال تحصیل"}
        reg_map = {0: "عادی", 1: "شهید", 2: "حکمت"}
        center_map = {1: "مرکز", 2: "گلستان", 3: "صدرا"}

        def to_jdate(d):
            if not d:
                return ""
            return jdatetime.datetime.fromgregorian(datetime=d).strftime("%Y/%m/%d %H:%M")

        def to_jdate_date(dd) -> str:
            if not dd:
                return ""
            return jdatetime.date.fromgregorian(date=dd).strftime("%Y/%m/%d")

        # Normalize phone to 0xxxxxxxxxx
        phone = s.phone
        if isinstance(phone, str) and phone.startswith("+98"):
            rest = phone[3:]
            if rest:
                phone = "0" + rest

        return [
            s.student_id,
            s.counter,
            s.first_name,
            s.last_name,
            s.national_code,
            phone,
            to_jdate_date(s.birth_date),
            gender_map.get(s.gender, ""),
            edu_map.get(s.education_status, ""),
            reg_map.get(s.registration_status, ""),
            center_map.get(s.center, ""),
            s.grade_level,
            ("مدرسه‌ای" if s.school_type == "school" else "عادی"),
            s.school_code or "",
            to_jdate(s.created_at),
            to_jdate(s.updated_at),
            (s.allocation_status or ""),
        ]
