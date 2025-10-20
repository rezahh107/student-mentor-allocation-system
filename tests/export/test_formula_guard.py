from __future__ import annotations

import csv
from dataclasses import replace
from datetime import datetime, timezone

from openpyxl import load_workbook

from sma.phase6_import_to_sabt.export_writer import EXPORT_COLUMNS
from sma.phase6_import_to_sabt.models import ExportFilters, ExportOptions, ExportSnapshot
from tests.export.helpers import build_exporter, make_row


def _assert_prefixed_once(value: str) -> None:
    assert value.startswith("'"), f"Formula prefix missing: {value!r}"
    assert not value.startswith("''"), f"Formula prefixed twice: {value!r}"


def test_formula_like_values_are_prefixed_once(tmp_path) -> None:
    base_time = datetime(2024, 3, 21, 8, 0, tzinfo=timezone.utc)
    raw = make_row(idx=3)
    row = replace(
        raw,
        first_name="=SUM(1,1)",
        last_name="+Value",
        mentor_name="-Danger",
        mentor_id="@lookup",
        mentor_mobile="\t09120000000",
    )

    exporter_csv = build_exporter(tmp_path / "csv", [row])
    manifest_csv = exporter_csv.run(
        filters=ExportFilters(year=1402, center=1),
        options=ExportOptions(output_format="csv"),
        snapshot=ExportSnapshot(marker="snap", created_at=base_time),
        clock_now=base_time,
    )

    csv_path = tmp_path / "csv" / manifest_csv.files[0].name
    with csv_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        data = next(reader)
    for column in ("first_name", "last_name", "mentor_name", "mentor_id"):
        _assert_prefixed_once(data[column])

    exporter_xlsx = build_exporter(tmp_path / "xlsx", [row])
    manifest_xlsx = exporter_xlsx.run(
        filters=ExportFilters(year=1402, center=1),
        options=ExportOptions(output_format="xlsx"),
        snapshot=ExportSnapshot(marker="snap", created_at=base_time),
        clock_now=base_time,
    )

    xlsx_path = tmp_path / "xlsx" / manifest_xlsx.files[0].name
    workbook = load_workbook(xlsx_path, read_only=True)
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(min_row=2, max_row=2, values_only=True)
        row_values = next(rows_iter)
        mapped = dict(zip(EXPORT_COLUMNS, row_values))
        for column in ("first_name", "last_name", "mentor_name", "mentor_id"):
            value = str(mapped[column])
            _assert_prefixed_once(value)
    finally:
        workbook.close()
