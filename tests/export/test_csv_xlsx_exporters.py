from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from src.infrastructure.persistence.models import AllocationRecord, Base
from src.tools.export.csv_exporter import export_allocations_to_csv
from src.tools.export.xlsx_exporter import export_allocations_to_xlsx


GOLDEN_DIR = Path(__file__).resolve().parents[1] / "golden" / "export"


@pytest.fixture()
def db_session(tmp_path):
    db_path = tmp_path / "export.sqlite"
    engine = create_engine(f"sqlite:///{db_path}", future=True)
    Base.metadata.create_all(engine, tables=[AllocationRecord.__table__])
    Session = sessionmaker(bind=engine, expire_on_commit=False, future=True)
    session = Session()
    yield session
    session.close()


def _add_sample_record(session) -> None:
    record = AllocationRecord(
        allocation_id=1,
        allocation_code="کد-۰۰۱",
        year_code="1402",
        student_id="۰۰۱۲۳۴۵۶۷۸",
        mentor_id=42,
        idempotency_key="abc",
        request_id="req-1",
        status="CONFIRMED",
        policy_code="=ALERT",
        metadata_json=None,
    )
    record.created_at = datetime(2024, 1, 1, tzinfo=timezone.utc)
    session.add(record)
    session.commit()


def test_csv_export_with_bom_and_crlf(db_session, tmp_path):
    _add_sample_record(db_session)
    output = tmp_path / "alloc_bom.csv"
    export_allocations_to_csv(
        session=db_session,
        output=output,
        bom=True,
        crlf=True,
        chunk_size=1,
    )
    expected = (GOLDEN_DIR / "allocations_bom.csv").read_bytes()
    assert output.read_bytes() == expected


def test_csv_export_without_bom(db_session, tmp_path):
    _add_sample_record(db_session)
    output = tmp_path / "alloc_no_bom.csv"
    export_allocations_to_csv(
        session=db_session,
        output=output,
        bom=False,
        crlf=False,
        chunk_size=1,
    )
    expected = (GOLDEN_DIR / "allocations_no_bom.csv").read_bytes()
    assert output.read_bytes() == expected


def test_xlsx_export_preserves_persian_digits(db_session, tmp_path):
    _add_sample_record(db_session)
    output = tmp_path / "alloc.xlsx"
    export_allocations_to_xlsx(session=db_session, output=output, chunk_size=1)
    workbook = load_workbook(output, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0][0] == "شناسه تخصیص"
    assert rows[1][2] == "1402"
    assert rows[1][3] == "۰۰۱۲۳۴۵۶۷۸"
    assert rows[1][6] == "'=ALERT"
