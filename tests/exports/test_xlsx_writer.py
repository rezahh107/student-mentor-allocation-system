from pathlib import Path

from openpyxl import load_workbook

from sma.phase6_import_to_sabt.xlsx.writer import XLSXStreamWriter


def _row(**kwargs):
    defaults = {
        "national_id": "001",
        "counter": "140257300",
        "first_name": "علی",
        "last_name": "رضایی",
        "gender": 0,
        "mobile": "۰۹۱۲۳۴۵۶۷۸۹",
        "reg_center": 1,
        "reg_status": 1,
        "group_code": 5,
        "student_type": 0,
        "school_code": 123,
        "mentor_id": "m-1",
        "mentor_name": "=SUM(A1)",
        "mentor_mobile": "09120000000",
        "allocation_date": "2023-01-01T00:00:00Z",
        "year_code": "1402",
    }
    defaults.update(kwargs)
    return defaults


def test_multi_sheet_chunking_stable_sort(tmp_path: Path) -> None:
    writer = XLSXStreamWriter(chunk_size=2)
    rows = [
        _row(national_id="002", school_code=321, reg_center=2),
        _row(national_id="003", school_code=111, reg_center=1),
        _row(national_id="001", school_code=222, reg_center=1),
    ]
    output = tmp_path / "export.xlsx"
    artifact = writer.write(rows, output)
    assert list(artifact.row_counts.values()) == [2, 1]
    wb = load_workbook(output, read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames
        assert sheet_names == ["Sheet_001", "Sheet_002"]
        first_sheet = wb[sheet_names[0]]
        values = [row[0].value for row in first_sheet.iter_rows(min_row=2)]
        assert values == ["003", "001"]
    finally:
        wb.close()


def test_sensitive_as_text_and_formula_guard(tmp_path: Path) -> None:
    writer = XLSXStreamWriter(chunk_size=10)
    output = tmp_path / "sensitive.xlsx"
    rows = [_row(national_id="۰۰۱", counter="۱۴۰۲۵۷۳۰۰", mentor_name="=2+2")]
    artifact = writer.write(rows, output)
    wb = load_workbook(output, read_only=True, data_only=False)
    try:
        sheet = wb.active
        cells = list(sheet.iter_rows(min_row=2, max_row=2))[0]
        national_id_cell = cells[0]
        mentor_name_cell = cells[12]
        assert national_id_cell.number_format == "@"
        assert str(national_id_cell.value).startswith("001")
        assert str(mentor_name_cell.value).startswith("'")
    finally:
        wb.close()
    assert artifact.excel_safety["sensitive_text"]
