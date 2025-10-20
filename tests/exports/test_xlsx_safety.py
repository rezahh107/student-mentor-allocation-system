from __future__ import annotations

from dataclasses import replace

import openpyxl
from openpyxl.styles import numbers

from sma.phase6_import_to_sabt.api import HMACSignedURLProvider, create_export_api
from sma.phase6_import_to_sabt.compat import TestClient
from sma.phase7_release.deploy import ReadinessGate

from tests.export.helpers import build_job_runner, make_row


def _build_export_api(tmp_path, rows):
    runner, metrics = build_job_runner(tmp_path, rows)
    signer = HMACSignedURLProvider(secret="secret")
    gate = ReadinessGate(clock=lambda: 0.0)
    gate.record_cache_warm()
    gate.record_dependency(name="redis", healthy=True)
    gate.record_dependency(name="database", healthy=True)
    app = create_export_api(
        runner=runner,
        signer=signer,
        metrics=metrics,
        logger=runner.logger,
        readiness_gate=gate,
    )
    return app, runner, metrics


def test_sensitive_as_text_and_formula_guard(tmp_path) -> None:
    """XLSX exports coerce sensitive columns to text and guard formulas."""

    sample = replace(
        make_row(idx=2),
        first_name="=SUM(A1:A2)",
        mentor_id="=1337",
        mentor_name="-1",
        national_id="۱۲۳۴۵۶۷۸۹۰",
        mobile="۰۹۱۲۳۴۵۶۷۸۸",
    )
    app, runner, metrics = _build_export_api(tmp_path, [sample])
    client = TestClient(app)

    response = client.post(
        "/exports",
        json={"year": 1402, "center": 1},
        headers={"Idempotency-Key": "idem-xlsx", "X-Role": "ADMIN"},
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    runner.await_completion(job_id)

    status = client.get(f"/exports/{job_id}")
    assert status.status_code == 200
    payload = status.json()

    manifest = payload["manifest"]
    assert manifest["format"] == "xlsx"
    safety = manifest["excel_safety"]
    assert safety["formula_guard"]
    assert set(safety["sensitive_text"]).issuperset({"national_id", "counter", "mobile", "mentor_id", "school_code"})

    file_name = manifest["files"][0]["name"]
    workbook = openpyxl.load_workbook(tmp_path / file_name, read_only=True)
    sheet = workbook.active
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    row_cells = next(sheet.iter_rows(min_row=2, max_row=2, values_only=False))
    header_map = {name: idx for idx, name in enumerate(header)}

    for column in ("national_id", "counter", "mobile", "mentor_id", "school_code"):
        cell = row_cells[header_map[column]]
        assert cell.number_format == numbers.FORMAT_TEXT
        assert cell.data_type == "s"

    first_name_cell = row_cells[header_map["first_name"]]
    assert isinstance(first_name_cell.value, str)
    assert first_name_cell.value.startswith("'=")

    rows_metric = metrics.rows_total.collect()[0].samples
    assert any(sample.labels.get("format") == "xlsx" and sample.value > 0 for sample in rows_metric)
    workbook.close()
