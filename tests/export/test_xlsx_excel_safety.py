from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import numbers

from src.phase6_import_to_sabt.xlsx.constants import SENSITIVE_COLUMNS
from src.phase6_import_to_sabt.xlsx.metrics import build_import_export_metrics
from src.phase6_import_to_sabt.xlsx.writer import XLSXStreamWriter


def test_formula_guard_and_sensitive_as_text(cleanup_fixtures) -> None:
    writer = XLSXStreamWriter(chunk_size=8)
    output_dir = cleanup_fixtures.base_dir / "exports"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "students.xlsx"
    metrics = build_import_export_metrics(cleanup_fixtures.registry)

    rows = [
        {
            "national_id": "۰۰۱۲۳۴۵۶۷۸",
            "counter": "9900123456",
            "first_name": "=SUM(A1:A2)",
            "last_name": "محمدی",
            "gender": 1,
            "mobile": "۰۹۱۲۳۴۵۶۷۸‌۹",
            "reg_center": 2,
            "reg_status": 1,
            "group_code": 14,
            "student_type": 3,
            "school_code": "123456",
            "mentor_id": "۴۵۶۷۸",
            "mentor_name": "لیلا",
            "mentor_mobile": "۰۹۳۵۶۷۸۹۰۱۲",
            "allocation_date": "1402-01-01",
            "year_code": "02",
        }
    ]

    artifact = writer.write(
        rows,
        output_path,
        metrics=metrics,
        format_label="xlsx",
        sleeper=lambda _: None,
    )
    debug_context = cleanup_fixtures.context(export_path=str(output_path))
    assert artifact.excel_safety["formula_guard"], debug_context
    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1))
    headers = {cell.value: index for index, cell in enumerate(header_row, start=1)}

    row_index = 2
    for column in SENSITIVE_COLUMNS:
        idx = headers[column]
        cell = sheet.cell(row=row_index, column=idx)
        assert cell.number_format == numbers.FORMAT_TEXT, debug_context
        assert cell.data_type == "s", debug_context
        assert not str(cell.value).startswith("="), debug_context

    first_name_cell = sheet.cell(row=row_index, column=headers["first_name"])
    assert str(first_name_cell.value).startswith("'=SUM"), debug_context
    mobile_cell = sheet.cell(row=row_index, column=headers["mobile"])
    assert mobile_cell.value == "09123456789", debug_context
    mentor_cell = sheet.cell(row=row_index, column=headers["mentor_id"])
    assert mentor_cell.number_format == numbers.FORMAT_TEXT, debug_context
    assert Path(artifact.path) == output_path, debug_context
