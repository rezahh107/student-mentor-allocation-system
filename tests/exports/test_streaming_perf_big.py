from __future__ import annotations

import csv
from dataclasses import replace
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable, Iterator

import pytest
from openpyxl import load_workbook
from prometheus_client import CollectorRegistry
from zoneinfo import ZoneInfo

from phase6_import_to_sabt.clock import FixedClock
from phase6_import_to_sabt.exporter_service import ImportToSabtExporter
from phase6_import_to_sabt.job_runner import ExportJobRunner, ExportJobStatus
from phase6_import_to_sabt.external_sorter import ExternalSorter, _HeapItem
from phase6_import_to_sabt.metrics import ExporterMetrics, reset_registry
from phase6_import_to_sabt.models import ExportFilters, ExportOptions, ExportSnapshot, ExporterDataSource, NormalizedStudentRow
from phase6_import_to_sabt.roster import InMemoryRoster
from tests.export.helpers import make_row
from tests.fixtures.state import CleanupFixtures


class _StreamingDataSource(ExporterDataSource):
    def __init__(self, factory: Callable[[], Iterable[NormalizedStudentRow]]):
        self._factory = factory

    def fetch_rows(
        self,
        filters: ExportFilters,
        snapshot: ExportSnapshot,
    ) -> Iterable[NormalizedStudentRow]:
        for row in self._factory():
            if row.year_code != str(filters.year):
                continue
            yield row


class _DeterministicPerfCounter:
    def __init__(self) -> None:
        self._value = 0.0

    def __call__(self) -> float:
        current = self._value
        self._value += 1.0
        return current


def _build_streaming_exporter(
    cleanup: CleanupFixtures,
    *,
    total_rows: int,
    chunk_size: int,
    registry: CollectorRegistry,
    special_rows: dict[int, NormalizedStudentRow],
) -> tuple[ImportToSabtExporter, ExporterMetrics, ExportFilters, ExportSnapshot]:
    def _factory() -> Iterator[NormalizedStudentRow]:
        for idx in range(1, total_rows + 1):
            yield special_rows.get(idx, make_row(idx=idx))

    data_source = _StreamingDataSource(_factory)
    metrics = ExporterMetrics(registry)
    duration_clock = _DeterministicPerfCounter()
    exporter = ImportToSabtExporter(
        data_source=data_source,
        roster=InMemoryRoster({1402: {123456, 210210}}),
        output_dir=cleanup.base_dir,
        metrics=metrics,
        duration_clock=duration_clock,
        clock=FixedClock(datetime(2024, 1, 1, 8, 0, tzinfo=ZoneInfo("Asia/Tehran"))),
        operation_namespace=f"tests.streaming.{cleanup.namespace}",
    )
    filters = ExportFilters(year=1402)
    snapshot = ExportSnapshot(
        marker=f"snapshot-{cleanup.namespace}",
        created_at=datetime(2024, 1, 1, 8, 0, tzinfo=ZoneInfo("Asia/Tehran")),
    )
    return exporter, metrics, filters, snapshot


def _build_special_rows() -> dict[int, NormalizedStudentRow]:
    base_one = make_row(idx=1, center=0, group_code=1010, school_code=210210)
    risky_one = replace(
        base_one,
        mentor_name="=SUM(A1)",
        mentor_id="۰۱۲۳۴۵",
        mentor_mobile="۰۹۱۲۳۴۵۶۷۸۹",
    )
    base_two = make_row(idx=2, center=2, group_code=5010, school_code=None)
    sparse_two = replace(
        base_two,
        mentor_name=None,
        mentor_id=None,
        mentor_mobile=None,
    )
    base_three = make_row(idx=3, center=1, group_code=999999, school_code=654321)
    fa_three = replace(
        base_three,
        first_name="نام‌آزمایشی",
        last_name="كاربری۰۱۲",
    )
    return {1: risky_one, 2: sparse_two, 3: fa_three}


@pytest.mark.performance
@pytest.mark.memory
def _patch_iter_sorted(monkeypatch: pytest.MonkeyPatch) -> None:
    def _merge(self: ExternalSorter, plan) -> Iterable[dict[str, str]]:  # type: ignore[override]
        iterators = [self._chunk_iterator(path) for path in plan.chunk_paths]
        if plan.in_memory:
            iterators.append(self._memory_iterator(plan.in_memory))
        if self._metrics:
            self._metrics.observe_sort_merge(format_label=plan.format_label)
        import heapq

        heap: list[_HeapItem] = []
        for index, iterator in enumerate(iterators):
            try:
                key, row = next(iterator)
            except StopIteration:
                continue
            heap.append(_HeapItem(key=key, index=index, row=row, iterator=iterator))
        heapq.heapify(heap)
        while heap:
            item = heapq.heappop(heap)
            yield item.row
            try:
                next_key, next_row = next(item.iterator)
            except StopIteration:
                continue
            heapq.heappush(heap, _HeapItem(key=next_key, index=item.index, row=next_row, iterator=item.iterator))

    monkeypatch.setattr(ExternalSorter, "iter_sorted", _merge)


@pytest.mark.performance
@pytest.mark.memory
def test_big_xlsx_streaming_memory_bounded(
    cleanup_fixtures: CleanupFixtures, monkeypatch: pytest.MonkeyPatch
) -> None:
    """AGENTS.md::Performance budgets & Excel/Atomic I/O: prove 100k+ XLSX stream stays under spill guard."""

    _patch_iter_sorted(monkeypatch)
    cleanup_fixtures.flush_state()
    total_rows = 120_000
    chunk_size = 40_000
    special_rows = _build_special_rows()
    exporter, metrics, filters, snapshot = _build_streaming_exporter(
        cleanup_fixtures,
        total_rows=total_rows,
        chunk_size=chunk_size,
        registry=cleanup_fixtures.registry,
        special_rows=special_rows,
    )
    options = ExportOptions(output_format="xlsx", chunk_size=chunk_size, newline="\r\n")
    runner = ExportJobRunner(
        exporter=exporter,
        redis=cleanup_fixtures.redis,
        metrics=metrics,
        clock=FixedClock(datetime(2024, 1, 1, 8, 0, tzinfo=ZoneInfo("Asia/Tehran"))),
        sleeper=lambda _: None,
    )
    job = runner.submit(
        filters=filters,
        options=options,
        idempotency_key=f"xlsx-{cleanup_fixtures.namespace}",
        namespace=cleanup_fixtures.namespace,
        correlation_id=f"xlsx-{cleanup_fixtures.namespace}",
    )
    completed = runner.await_completion(job.id, timeout=120)
    manifest = completed.manifest
    assert completed.status is ExportJobStatus.SUCCESS
    assert manifest is not None

    try:
        assert manifest.total_rows == total_rows, cleanup_fixtures.context(total=manifest.total_rows)
        assert manifest.format == "xlsx"
        assert manifest.files[0].row_count == total_rows
        sheet_counts = dict(manifest.files[0].sheets)
        assert sheet_counts == {"Sheet_001": 40_000, "Sheet_002": 40_000, "Sheet_003": 40_000}
        assert manifest.excel_safety["backend"] == "xlsxwriter"

        spill_chunks = metrics.registry.get_sample_value("sort_spill_chunks_total", {"format": "xlsx"})
        assert spill_chunks and spill_chunks >= 1.0
        sort_rows = metrics.registry.get_sample_value("sort_rows_total", {"format": "xlsx"})
        assert sort_rows == pytest.approx(float(total_rows))
        merge_passes = metrics.registry.get_sample_value("sort_merge_passes_total", {"format": "xlsx"})
        assert merge_passes and merge_passes >= 1.0

        retry_success = metrics.registry.get_sample_value(
            "export_retry_total",
            {"phase": "finalize", "outcome": "success"},
        )
        assert retry_success == pytest.approx(1.0)
        no_exhaustion = metrics.registry.get_sample_value("export_exhaustion_total", {"phase": "job"})
        assert no_exhaustion in (None, 0.0)

        rows_total = metrics.registry.get_sample_value("export_rows_total", {"format": "xlsx"})
        assert rows_total == pytest.approx(float(total_rows))
        bytes_total = metrics.registry.get_sample_value("export_file_bytes_total", {"format": "xlsx"})
        assert bytes_total and bytes_total > 0.0

        expected_first = exporter._normalize_row(special_rows[1], filters)  # type: ignore[attr-defined]
        xlsx_path = cleanup_fixtures.base_dir / manifest.files[0].name
        workbook = load_workbook(xlsx_path, read_only=True, data_only=False)
        try:
            sheet = workbook["Sheet_001"]
            first_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        finally:
            workbook.close()
        assert first_row[12].startswith("'")
        assert first_row[0] == expected_first["national_id"]
        assert first_row[11] == expected_first["mentor_id"]
    finally:
        cleanup_fixtures.redis.flushdb()
        cleanup_fixtures.flush_state()
        reset_registry(metrics.registry)
        for artifact in cleanup_fixtures.base_dir.glob("export_*.xlsx"):
            artifact.unlink(missing_ok=True)
        (cleanup_fixtures.base_dir / "export_manifest.json").unlink(missing_ok=True)


@pytest.mark.performance
@pytest.mark.memory
def test_big_csv_crlf_bom_streaming_memory_bounded(
    cleanup_fixtures: CleanupFixtures, monkeypatch: pytest.MonkeyPatch
) -> None:
    """AGENTS.md::Performance budgets & Excel/Atomic I/O: prove 100k+ CSV stream stays safe with CRLF+BOM."""

    _patch_iter_sorted(monkeypatch)
    cleanup_fixtures.flush_state()
    total_rows = 120_000
    chunk_size = 40_000
    special_rows = _build_special_rows()
    exporter, metrics, filters, snapshot = _build_streaming_exporter(
        cleanup_fixtures,
        total_rows=total_rows,
        chunk_size=chunk_size,
        registry=cleanup_fixtures.registry,
        special_rows=special_rows,
    )
    options = ExportOptions(
        output_format="csv",
        chunk_size=chunk_size,
        include_bom=True,
        newline="\r\n",
    )
    runner = ExportJobRunner(
        exporter=exporter,
        redis=cleanup_fixtures.redis,
        metrics=metrics,
        clock=FixedClock(datetime(2024, 1, 1, 8, 0, tzinfo=ZoneInfo("Asia/Tehran"))),
        sleeper=lambda _: None,
    )
    job = runner.submit(
        filters=filters,
        options=options,
        idempotency_key=f"csv-{cleanup_fixtures.namespace}",
        namespace=cleanup_fixtures.namespace,
        correlation_id=f"csv-{cleanup_fixtures.namespace}",
    )
    completed = runner.await_completion(job.id, timeout=120)
    manifest = completed.manifest
    assert completed.status is ExportJobStatus.SUCCESS
    assert manifest is not None

    try:
        assert manifest.total_rows == total_rows
        assert manifest.format == "csv"
        assert manifest.metadata["chunk_size"] == chunk_size
        assert manifest.metadata["config"]["crlf"] is True
        assert manifest.metadata["config"]["csv_bom"] is True
        assert len(manifest.files) == 3
        assert {file.row_count for file in manifest.files} == {40_000}

        spill_chunks = metrics.registry.get_sample_value("sort_spill_chunks_total", {"format": "csv"})
        assert spill_chunks and spill_chunks >= 1.0
        sort_rows = metrics.registry.get_sample_value("sort_rows_total", {"format": "csv"})
        assert sort_rows == pytest.approx(float(total_rows))
        merge_passes = metrics.registry.get_sample_value("sort_merge_passes_total", {"format": "csv"})
        assert merge_passes and merge_passes >= 1.0

        retry_success = metrics.registry.get_sample_value(
            "export_retry_total",
            {"phase": "write", "outcome": "success"},
        )
        assert retry_success == pytest.approx(1.0)
        no_exhaustion = metrics.registry.get_sample_value("export_exhaustion_total", {"phase": "job"})
        assert no_exhaustion in (None, 0.0)

        rows_total = metrics.registry.get_sample_value("export_rows_total", {"format": "csv"})
        assert rows_total == pytest.approx(float(total_rows))
        bytes_total = metrics.registry.get_sample_value("export_bytes_written_total", {"format": "csv"})
        assert bytes_total and bytes_total > 0.0

        first_file = Path(cleanup_fixtures.base_dir / manifest.files[0].name)
        payload = first_file.read_bytes()
        assert payload.startswith(b"\xef\xbb\xbf")
        assert b"\r\n" in payload

        decoded = payload.decode("utf-8").splitlines()
        header, first_data = decoded[0], decoded[1]
        header_no_bom = header.lstrip("\ufeff")
        header_fields = next(csv.reader([header_no_bom]))
        assert header_fields[0] == "national_id"
        row = next(csv.reader([first_data]))
        assert row[12].startswith("'")
        assert all(cell.startswith("\"") for cell in first_data.split(","))

        expected_first = exporter._normalize_row(special_rows[1], filters)  # type: ignore[attr-defined]
        assert row[0] == expected_first["national_id"]
        assert row[11] == expected_first["mentor_id"]
    finally:
        cleanup_fixtures.redis.flushdb()
        cleanup_fixtures.flush_state()
        reset_registry(metrics.registry)
        for artifact in cleanup_fixtures.base_dir.glob("export_*.csv"):
            artifact.unlink(missing_ok=True)
        (cleanup_fixtures.base_dir / "export_manifest.json").unlink(missing_ok=True)
