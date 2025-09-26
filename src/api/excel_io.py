"""Streaming Persian-friendly Excel and CSV utilities."""
from __future__ import annotations

import contextlib
import csv
import gc
import io
import tempfile
import unicodedata
import zipfile
import xml.etree.ElementTree as ET
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import BinaryIO, Iterable, Iterator, Sequence

from src.api.patterns import formula_prefixes, zero_width_pattern
from src.core.normalize import normalize_digits

try:  # pragma: no cover - optional dependency
    from openpyxl import Workbook
    from openpyxl.reader.excel import load_workbook
except Exception:  # pragma: no cover - allow CSV-only usage in environments without extras
    Workbook = None  # type: ignore[assignment]
    load_workbook = None  # type: ignore[assignment]

HAS_OPENPYXL = Workbook is not None

FORMULA_PREFIXES = formula_prefixes()
ZERO_WIDTH_RE = zero_width_pattern()
DEFAULT_MEMORY_LIMIT = 32 * 1024 * 1024
DEFAULT_CREATED_AT = datetime(2024, 1, 1, 0, 0, 0)


class ExcelMemoryError(RuntimeError):
    """Raised when streaming exceeds the configured memory budget."""


class _BoundedBuffer:
    """A streaming buffer that tracks whether the size budget was exceeded."""

    def __init__(self, *, limit: int) -> None:
        self._limit = max(1, limit)
        self._file = tempfile.SpooledTemporaryFile(max_size=self._limit)
        self._overflowed = False

    def write(self, data: object) -> int:
        if not isinstance(data, (bytes, bytearray, memoryview)):
            raise TypeError("expected bytes-like object")
        chunk = bytes(data)
        if self._file.tell() + len(chunk) > self._limit:
            self._overflowed = True
        return self._file.write(chunk)

    def getvalue(self) -> bytes:
        self._file.seek(0)
        return self._file.read()

    def seek(self, offset: int, whence: int = io.SEEK_SET) -> int:
        return self._file.seek(offset, whence)

    def tell(self) -> int:
        return self._file.tell()

    def read(self, size: int = -1) -> bytes:
        return self._file.read(size)

    def flush(self) -> None:
        self._file.flush()

    def close(self) -> None:
        self._file.close()

    @property
    def overflowed(self) -> bool:
        return self._overflowed


@contextmanager
def _bounded_buffer(limit: int) -> Iterator[_BoundedBuffer]:
    buffer = _BoundedBuffer(limit=limit)
    try:
        yield buffer
    finally:
        buffer.close()


@dataclass(slots=True)
class ExcelRow:
    """Represents a normalized row returned from Excel import."""

    cells: Sequence[str]


def _normalize_text(value: str) -> str:
    text = unicodedata.normalize("NFKC", value)
    text = normalize_digits(text)
    text = ZERO_WIDTH_RE.sub("", text)
    return text.strip()


def sanitize_cell(value: object) -> str:
    """Normalize and sanitize a single cell value for textual output."""

    if value is None:
        return ""
    text = _normalize_text(str(value))
    if text.startswith(FORMULA_PREFIXES):
        return "'" + text
    return text


def _prepare_xlsx_cell(value: object) -> object:
    """Prepare a value for XLSX output while preserving numbers/dates."""

    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, float, Decimal)):
        return value
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.astimezone(timezone.utc).replace(tzinfo=None)
        return value
    if isinstance(value, date):
        return value
    text = _normalize_text(str(value))
    if text.startswith(FORMULA_PREFIXES):
        text = "'" + text
    return text


def iter_csv_rows(stream: io.TextIOBase) -> Iterator[ExcelRow]:
    """Stream rows from a CSV file while normalizing content."""

    reader = csv.reader(stream)
    for raw_row in reader:
        yield ExcelRow(cells=[sanitize_cell(cell) for cell in raw_row])


def write_csv(rows: Iterable[Sequence[object]], *, stream: BinaryIO) -> None:
    """Write rows into a CSV stream using UTF-8 with BOM for Persian compatibility."""

    text_stream = io.TextIOWrapper(stream, encoding="utf-8-sig", newline="", write_through=True)
    writer = csv.writer(text_stream, quoting=csv.QUOTE_MINIMAL)
    for row in rows:
        writer.writerow([sanitize_cell(cell) for cell in row])
    text_stream.flush()
    text_stream.detach()


def iter_xlsx_rows(stream: BinaryIO, *, sheet: str | None = None) -> Iterator[ExcelRow]:
    """Stream rows from an XLSX workbook using openpyxl's read-only mode."""

    if load_workbook is None:  # pragma: no cover - optional dependency guard
        raise RuntimeError("Install with excel extra to read XLSX files")
    workbook = load_workbook(stream, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet] if sheet else workbook.active
        for row in worksheet.iter_rows(values_only=True):
            yield ExcelRow(cells=[sanitize_cell(cell) for cell in row])
    finally:
        workbook.close()


def write_xlsx(
    rows: Iterable[Sequence[object]],
    *,
    sheet_name: str = "Sheet1",
    memory_limit_bytes: int = DEFAULT_MEMORY_LIMIT,
) -> bytes:
    """Generate an XLSX workbook in memory using streaming writers."""

    if Workbook is None:  # pragma: no cover - optional dependency guard
        raise RuntimeError("Install with excel extra to write XLSX files")
    workbook = Workbook(write_only=True)
    workbook.iso_dates = True
    workbook.properties.creator = "allocation-api"
    workbook.properties.lastModifiedBy = "allocation-api"
    workbook.properties.created = DEFAULT_CREATED_AT
    workbook.properties.modified = DEFAULT_CREATED_AT
    worksheet = workbook.create_sheet(title=sheet_name)
    for row in rows:
        worksheet.append([_prepare_xlsx_cell(cell) for cell in row])
    with _bounded_buffer(memory_limit_bytes) as raw_buffer:
        workbook.save(raw_buffer)
        if raw_buffer.overflowed:
            _close_write_only_handles(workbook)
            workbook.close()
            gc.collect()
            raise ExcelMemoryError("حجم خروجی اکسل از حد مجاز فراتر رفت")
        payload = raw_buffer.getvalue()
    _close_write_only_handles(workbook)
    workbook.close()
    gc.collect()
    return _repack_xlsx(payload, memory_limit_bytes)


def _repack_xlsx(payload: bytes, memory_limit_bytes: int) -> bytes:
    """Repack XLSX bytes with deterministic ordering and timestamps."""

    with _bounded_buffer(memory_limit_bytes) as target:
        with zipfile.ZipFile(io.BytesIO(payload), "r") as source:
            with zipfile.ZipFile(target, "w", compression=zipfile.ZIP_DEFLATED) as sink:
                for info in sorted(source.infolist(), key=lambda entry: entry.filename):
                    data = source.read(info.filename)
                    if info.filename == "docProps/core.xml":
                        data = _normalise_core_props(data)
                    new_info = zipfile.ZipInfo(filename=info.filename, date_time=(2024, 1, 1, 0, 0, 0))
                    new_info.compress_type = zipfile.ZIP_DEFLATED
                    new_info.external_attr = info.external_attr
                    sink.writestr(new_info, data)
        if target.overflowed:
            raise ExcelMemoryError("حجم خروجی اکسل از حد مجاز فراتر رفت")
        return target.getvalue()


def _normalise_core_props(payload: bytes) -> bytes:
    """Overwrite core property timestamps for deterministic builds."""

    root = ET.fromstring(payload)
    ns = {
        "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
        "dcterms": "http://purl.org/dc/terms/",
    }
    constant = "2024-01-01T00:00:00Z"
    for tag in ("dcterms:created", "dcterms:modified"):
        element = root.find(tag, ns)
        if element is not None:
            element.text = constant
    return ET.tostring(root, encoding="utf-8", xml_declaration=False)
def _close_write_only_handles(workbook: "Workbook") -> None:
    """Best-effort cleanup for openpyxl write-only internals."""

    if not HAS_OPENPYXL:  # pragma: no cover - defensive guard
        return
    with contextlib.suppress(Exception):
        archive = getattr(workbook, "_archive", None)
        if archive is not None:
            archive.close()
    sheets = getattr(workbook, "_sheets", [])
    for sheet in sheets:
        writer = getattr(sheet, "_writer", None)
        if writer is not None:
            with contextlib.suppress(Exception):
                writer.close()
